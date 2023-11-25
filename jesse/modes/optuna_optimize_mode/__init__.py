import warnings
warnings.filterwarnings('ignore')
from timeloop import Timeloop
from datetime import timedelta
import logging
import os
from jesse import exceptions
isfile = os.path.isfile
join = os.path.join

from random import randint, choices
import pathlib
import matplotlib.pyplot as plt
import pickle
import shutil
import subprocess
import traceback
import webbrowser
import math
import ast
from jesse.routes import router
import random
import spectra
from PIL import Image
from optuna_fast_fanova import FanovaImportanceEvaluator
# import gc 
import click
from pyprobs import Probability as pr
import jesse.helpers as jh
import numpy as np
import optuna
import pkg_resources
import yaml
import joblib
import copy
from jesse.research import backtest, get_candles
#from .JoblilbStudy import JoblibStudy
from subprocess import * 
import pandas as pd 
import datetime
import glob 
from datetime import datetime,timedelta, date
from typing import Iterable
from openpyxl.chart.axis import DateAxis
import itertools
from itertools import chain
import threading
import time
import sys
from colorama import Fore, Back, Style
from time import sleep
import xlsxwriter 
from openpyxl.drawing.text import CharacterProperties
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import BarChart, Series, Reference, LineChart, ScatterChart
from openpyxl.chart.label import DataLabelList
from termcolor import colored
from pyfiglet import Figlet
from joblib import Parallel, delayed
from time import gmtime
from time import strftime
from openpyxl.drawing.text import (
    ParagraphProperties,
    CharacterProperties,
)
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string, get_column_letter
from jesse.services.redis import process_status, sync_publish
from jesse.services.failure import register_custom_exception_handler
import signal
import itertools
# import matplotlib
# matplotlib.use("svg")
# import matplotlib.pyplot as plt 

#Dashboard Command
# optuna-dashboard postgresql://jesse_user:password@localhost/optuna_db

joblib_study_finished_event = threading.Event()

class JoblibStudy:
    def __init__(self, **study_parameters):
        self.study_parameters = study_parameters
        self.study: optuna.study.Study = optuna.create_study(**study_parameters)
        
    def _optimize_study(self, func, n_trials, **optimize_parameters):
        study_parameters = copy.copy(self.study_parameters)
        study_parameters["study_name"] = self.study.study_name
        study_parameters["load_if_exists"] = True
        study = optuna.create_study(**study_parameters)
        study.sampler.reseed_rng()
        study.optimize(func, n_trials=n_trials, catch=(Exception,), gc_after_trial=True, **optimize_parameters)

    @staticmethod
    def _split_trials(n_trials, n_jobs):
        n_per_job, remaining = divmod(n_trials, n_jobs)
        for _ in range(n_jobs):
            yield n_per_job + (1 if remaining > 0 else 0)
            remaining -= 1
            
    def optimize(self, func, n_trials=1, n_jobs=-1, **optimize_parameters):
        if n_jobs == -1:
            n_jobs = joblib.cpu_count()
            joblib_study_finished_event.set()
            print('Optimization finished')
        if n_jobs == 1:
            self.study.optimize(func, n_trials=n_trials, gc_after_trial=True, **optimize_parameters)
            joblib_study_finished_event.set()
            print('Optimization finished')
        else:
            with joblib.Parallel(n_jobs, backend='loky', verbose=10, max_nbytes=None) as parallel:
                parallel(joblib.delayed(self._optimize_study)(func, n_trials=n_trials_i, **optimize_parameters)
                         for n_trials_i in self._split_trials(n_trials, n_jobs))

            print('Optimization finished')
            # releasing workers early rather than default 300 seconds
            #from joblib.externals.loky import get_reusable_executor
            #get_reusable_executor().shutdown(wait=True)
            joblib_study_finished_event.set()
            
    def set_user_attr(self, key: str, value):
        if isinstance(value, np.integer):
            value = int(value)
        elif isinstance(value, np.floating):
            value = float(value)
        elif isinstance(value, np.ndarray):
            value = value.tolist()
        self.study.set_user_attr(key, value)

    def __getattr__(self, name):
        if not name.startswith("_") and hasattr(self.study, name):
            return getattr(self.study, name)
        else:
            raise AttributeError(f"'{self.__class__.__name__}' object has no attribute '{name}'")
            
            
logger = logging.getLogger()
logger.addHandler(logging.FileHandler("jesse-optuna.log", mode="w"))

optuna.logging.enable_propagation()

from joblib import Memory 
cachedir = './temp_cache/'
memory = Memory(cachedir, verbose=0)


def clear_cache() -> None: 
    memory.clear(warn=False)

def clear_database() -> None:
    from jesse.services.env import ENV_VALUES as cfg_
    import psycopg2
    from psycopg2 import sql

    db_name = 'optuna_db'
    db_user = 'jesse_user'
    db_password = cfg_['POSTGRES_PASSWORD'] 
    db_host = cfg_['POSTGRES_HOST'] 
    db_port = cfg_['POSTGRES_PORT'] 

    # Connect to the default database (or another database that is not being deleted)
    conn = psycopg2.connect(
        dbname="postgres",  # default database
        user=db_user,
        password=db_password,
        host=db_host,
        port=db_port
    )
    conn.autocommit = True  # Required for database creation and dropping

    try:
        # Open a cursor to perform database operations
        cur = conn.cursor()

        # Drop the existing database
        cur.execute(sql.SQL("DROP DATABASE IF EXISTS {}").format(sql.Identifier(db_name)))

        # Create a new database
        cur.execute(sql.SQL("CREATE DATABASE {}").format(sql.Identifier(db_name)))

        print("Database recreated successfully.")

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close communication with the database
        if cur is not None:
            cur.close()
        if conn is not None:
            conn.close()
            

def create_config() -> None:
    validate_cwd()
    target_dirname = pathlib.Path().resolve()
    package_dir = pathlib.Path(__file__).resolve().parent
    shutil.copy2(f'{package_dir}/optuna_config.yml', f'{target_dirname}/optuna_config.yml')

def create_db() -> None:

    validate_cwd()
    import psycopg2
    from jesse.services.env import ENV_VALUES as cfg_
    db_name = 'optuna_db'
    # establishing the connection
    conn = psycopg2.connect(
        database="postgres", user='jesse_user', password=cfg_['POSTGRES_PASSWORD'], host=cfg_['POSTGRES_HOST'], port=cfg_['POSTGRES_PORT']
    )
    conn.autocommit = True

    # Creating a cursor object using the cursor() method
    cursor = conn.cursor()

    # Creating a database
    cursor.execute('CREATE DATABASE ' + str(db_name))
    print(f"Database {db_name} created successfully...")

    # Closing the connection
    conn.close()

def check_study_exists(study_name):
    import psycopg2
    try:
        from jesse.services.env import ENV_VALUES as cfg_
        db_name = 'optuna_db'
        user = 'jesse_user'
        password = cfg_['POSTGRES_PASSWORD']
        host = cfg_['POSTGRES_HOST']
        port = cfg_['POSTGRES_PORT']
        # Connect to the PostgreSQL database server
        conn = psycopg2.connect(dbname=db_name, user=user, password=password, host=host, port=port)

        cur = conn.cursor()

        cur.execute("SELECT COUNT(*) FROM studies WHERE study_name = %s", (study_name,))

        result = cur.fetchone()
        exists = result[0] > 0

        cur.close()
        conn.close()

        return exists

    except Exception as error:
        print(f"An error occurred: {error}")
        try:
            conn = psycopg2.connect(dbname=db_name, user=user, password=password, host=host, port=port)
            cur = conn.cursor()
            cur.execute("SELECT study_name FROM studies")
            all_studies = cur.fetchall()
            print("All study names in the database:")
            for study in all_studies:
                print(study[0])

            cur.close()
        except Exception as inner_error:
            print(f"Failed to fetch study names: {inner_error}")
        finally:
            if conn:
                conn.close()

        return False
        
def database_exists():
    import psycopg2
    from psycopg2 import sql
    conn = None
    exists = False
    dbname = 'optuna_db'
    user = 'jesse_user'
    from jesse.services.env import ENV_VALUES as cfg
    password = cfg['POSTGRES_PASSWORD']
    host = cfg['POSTGRES_HOST']
    port = cfg['POSTGRES_PORT']
    try:
        # Connect to the PostgreSQL server
        conn = psycopg2.connect(dbname='postgres', user=user, password=password, host=host, port=port)
        conn.autocommit = True

        # Create a cursor object
        cur = conn.cursor()

        # Check if the database exists
        cur.execute(sql.SQL("SELECT 1 FROM pg_database WHERE datname = %s;"), (dbname,))
        exists = cur.fetchone() is not None

        # Close the cursor and connection
        cur.close()
    except psycopg2.Error as e:
        print(f"Database check failed: {e}")
    finally:
        if conn is not None:
            conn.close()

    return exists
    
def generate_grid(hp):
    if hp['type'] == int:
        if 'step' not in hp:
            hp['step'] = 1
        return list(range(hp['min'], hp['max'] + 1, hp.get('step', 1)))
    elif hp['type'] == float:
        if 'step' not in hp:
            hp['step'] = 1
        return [hp['min'] + i * hp.get('step', 0.1) for i in range(int((hp['max'] - hp['min']) / hp.get('step', 0.1)) + 1)]
    elif hp['type'] == bool:
        return [True, False]
    elif hp['type'] == str:
        return hp['list']
    else:
        raise TypeError('Only int, bool, float, and str types are implemented for strategy parameters.')


def divide_date_range(start_date, finish_date, ratio=0.85):
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
    finish_date_obj = datetime.strptime(finish_date, '%Y-%m-%d')
    total_duration = finish_date_obj - start_date_obj
    first_period_duration = timedelta(days=total_duration.days * ratio)
    first_period_end = start_date_obj + first_period_duration
    first_period_end_rounded = first_period_end.replace(day=1)
    second_period_start = first_period_end_rounded

    return (
        [start_date, first_period_end_rounded.strftime('%Y-%m-%d'),
        second_period_start.strftime('%Y-%m-%d'), finish_date]
    )
    
def format_routes_to_dict(routes):
    formatted_routes = {"route": {}}
    for index, route in enumerate(routes):
        formatted_routes["route"][index] = {
            "exchange": route.get("exchange", ""),
            "symbol": route.get("symbol", ""),
            "timeframe": route.get("timeframe", ""),
        }
    return formatted_routes
    
def format_extra_routes_to_dict(routes):
    formatted_routes = {"extra_routes": {}}
    for index, route in enumerate(routes):
        formatted_routes["extra_routes"][index] = {
            "exchange": route.get("exchange", ""),
            "symbol": route.get("symbol", ""),
            "timeframe": route.get("timeframe", ""),
        }
    return formatted_routes

def is_port_in_use(port):
    import socket 
    import subprocess
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('localhost', port)) == 0
        
def reset_termination_signal():
    termination_file = "./storage/temp/termination_signal.txt"
    if os.path.exists(termination_file):
        os.remove(termination_file)
        
def signal_handler(signum, frame):
    reset_termination_signal()
    sys.exit(1)

signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)
    
def set_termination_signal():
    open("./storage/temp/termination_signal.txt", "w").close()
    

def run(
        continueExistingStudy: bool,
        debug_mode: bool, 
        config: dict, 
        routes: list, 
        extra_routes: list, 
        start_date: str, 
        finish_date: str, 
        optimal_total: int,
        optimizer: str,
        fitnessMetric1: str,
        fitnessMetric2: str,
        isMultivariate: bool,
        secondObjectiveDirection: str,
        consider_prior: bool,
        prior_weight: float,
        n_startup_trials: int,
        n_ei_candidates: int,
        group: bool,
        sigma: float,
        consider_pruned_trials: bool,
        population_size: int,
        crossover_prob: float,
        swapping_prob: float,
        qmc_type: str,
        scramble: bool,
        n_trials: int):
        
    validate_cwd()
    reset_termination_signal()
    db_exists = database_exists()
    if not db_exists:
        create_db()
        
    import jesse.helpers as jh
    from jesse.store import store 
        
    status_checker = Timeloop()
    @status_checker.job(interval=timedelta(seconds=1))
    def handle_time():
        if process_status() != 'started':
            set_termination_signal()
            joblib_study_finished_event.wait(100)
            raise exceptions.Termination
            
    status_checker.start()
    register_custom_exception_handler('optuna')
    store.app.set_session_id()
        
    dates_list = divide_date_range(start_date,finish_date) 

    from jesse.services.env import ENV_VALUES as _cfg_

    cfg = {
        "debug_mode": debug_mode,
        "type": config['exchange']['type'],
        "validation_interval": 90,
        "warm_up_candles": config['warm_up_candles'],
        "settlement_currency": "USD",
        "starting_balance": config['exchange']['balance'],
        "strategy_name": routes[0]['strategy'],
        "postgres_db_name": "optuna_db",
        "postgres_host": _cfg_['POSTGRES_HOST'],
        "postgres_password": _cfg_['POSTGRES_PASSWORD'],
        "postgres_port": _cfg_['POSTGRES_PORT'],
        "postgres_username": "jesse_user",
        "robust_test_iteration_count": 3,
        "robust_test_max": 6,
        "robust_test_min": 1.5,
        "dual_mode": secondObjectiveDirection,
        "fee": config['exchange']['fee'],
        "fitness-ratio1": fitnessMetric1,
        "fitness-ratio2": fitnessMetric2,
        "futures_leverage": config['exchange']['futures_leverage'],
        "futures_leverage_mode": config['exchange']['futures_leverage_mode'],
        "max_final_number_of_validation_results": 64,
        "multivariate": isMultivariate,
        "n_jobs": config['cpu_cores'],
        "n_trials": n_trials,
        "optimal-total": optimal_total,
        "optimizer": optimizer,
        "Interval_end": finish_date,
        "Interval_start": start_date,
        "pair-testing":{
            "finish_date": dates_list[3],
            "start_date": dates_list[0],
            },
        "timespan-testing": {
            "finish_date": dates_list[3], 
            "start_date": dates_list[2],
            },
        "timespan-train": {
            "finish_date": dates_list[1], 
            "start_date": dates_list[0],
            },
        }
        # Optimizer-specific settings
    optimizer_settings = {
        'TPESampler': {
            "multivariate" : isMultivariate,
            "consider_prior": consider_prior,
            "prior_weight": prior_weight,
            "n_startup_trials": n_startup_trials,
            "n_ei_candidates": n_ei_candidates,
            "group": group
        },
        'NSGAIISampler': {
            "population_size": population_size,
            "crossover_prob": crossover_prob,
            "swapping_prob": swapping_prob
        },
        'CmaEsSampler': {
            "sigma": sigma,  
            "n_startup_trials": n_startup_trials,
            "consider_pruned_trials": consider_pruned_trials 
        },
        'NSGAIIISampler': {
            "population_size": population_size,
            "crossover_prob": crossover_prob,
            "swapping_prob": swapping_prob
        },
        'MOTPESampler': {
            "consider_prior": consider_prior,
            "prior_weight": prior_weight,
            "n_startup_trials": n_startup_trials,
            "n_ei_candidates": n_ei_candidates,
        },
        'QMCSampler': {
            "qmc_type": qmc_type,
            "scramble": scramble
        },
        'GridSampler': {},
        'RandomSampler': {},
        'BruteForceSampler': {}
        # Add additional settings for other samplers if needed
    }

    # Add selected optimizer's settings to global config
    if optimizer in optimizer_settings:
        cfg[optimizer] = optimizer_settings[optimizer]

    # Additional settings based on multivariate mode
    if isMultivariate:
        cfg["secondObjectiveDirection"] = secondObjectiveDirection
        cfg['dual_mode'] = secondObjectiveDirection
        cfg['mode'] = 'multi'
    else:
        cfg['mode'] = 'single'
    
    routes_dict = format_routes_to_dict(routes)
    cfg.update(routes_dict)
    if extra_routes != {}:
        extra_routes_dict = format_extra_routes_to_dict(extra_routes)
        cfg.update(extra_routes_dict)
    start_date = cfg['timespan-train']['start_date']
    finish_date = cfg['timespan-testing']['finish_date']
    # {cfg['optimizer']}-{len(cfg['route'].items())} Pairs
    study_name = f"{routes[0]['strategy']}-{routes[0]['exchange']}-{routes[0]['symbol']}-{routes[0]['timeframe']}-{start_date}-{finish_date}"
    storage = f"postgresql://jesse_user:{cfg['postgres_password']}@{cfg['postgres_host']}:{cfg['postgres_port']}/optuna_db"
    #pruners defined with optimizer
    defined_pruner = None
    if cfg['optimizer'] == 'GridSampler': 
        StrategyClass = jh.get_strategy_class(cfg['strategy_name'])
        hp_dict = StrategyClass().hyperparameters()
        search_space = {hp['name']: generate_grid(hp) for hp in hp_dict}
        sampler = optuna.samplers.GridSampler(search_space)
    elif cfg['optimizer'] == 'BruteForceSampler':
        sampler = optuna.samplers.BruteForceSampler()
    elif cfg['optimizer'] == 'QMCSampler':
        sampler = optuna.samplers.QMCSampler(qmc_type=cfg[cfg['optimizer']]['qmc_type'], scramble=cfg[cfg['optimizer']]['scramble'])
        if cfg['mode'] == 'single': 
            defined_pruner = optuna.pruners.MedianPruner()
    elif cfg['optimizer'] == 'CmaEsSampler':
        sampler = optuna.samplers.CmaEsSampler(sigma0=cfg[cfg['optimizer']]['sigma'],n_startup_trials=cfg[cfg['optimizer']]['n_startup_trials'],consider_pruned_trials=cfg[cfg['optimizer']]['consider_pruned_trials'])
        if cfg['mode'] == 'single': 
            defined_pruner = optuna.pruners.PatientPruner()
    elif cfg['optimizer'] == 'RandomSampler':
        sampler = optuna.samplers.RandomSampler()
        if cfg['mode'] == 'single': 
            defined_pruner = optuna.pruners.MedianPruner()
    elif cfg['optimizer'] == 'NSGAIISampler': 
        sampler = optuna.samplers.NSGAIISampler(population_size=cfg[cfg['optimizer']]['population_size'],
                                                crossover_prob=cfg[cfg['optimizer']]['crossover_prob'], swapping_prob=cfg[cfg['optimizer']]['swapping_prob'],constraints_func=constraints_function)
        if cfg['mode'] == 'single': 
            defined_pruner = optuna.pruners.SuccessiveHalvingPruner()
    elif cfg['optimizer'] == 'NSGAIIISampler': 
        sampler = optuna.samplers.NSGAIIISampler(population_size=cfg[cfg['optimizer']]['population_size'],
                                                crossover_prob=cfg[cfg['optimizer']]['crossover_prob'], swapping_prob=cfg[cfg['optimizer']]['swapping_prob'])
        if cfg['mode'] == 'single': 
            defined_pruner = optuna.pruners.SuccessiveHalvingPruner()
    elif cfg['optimizer'] == 'TPESampler':
        sampler = optuna.samplers.TPESampler(multivariate=cfg[cfg['optimizer']]['multivariate'], consider_prior= cfg[cfg['optimizer']]['consider_prior'],n_startup_trials=cfg[cfg['optimizer']]['n_startup_trials'], n_ei_candidates=cfg[cfg['optimizer']]['n_ei_candidates'],
                                                prior_weight= cfg[cfg['optimizer']]['prior_weight'], group = cfg[cfg['optimizer']]['group'], constraints_func=constraints_function)
        if cfg['mode'] == 'single': 
            defined_pruner=optuna.pruners.HyperbandPruner()
    elif cfg['optimizer'] == 'MOTPESampler':
        sampler = optuna.samplers.MOTPESampler(consider_prior= cfg[cfg['optimizer']]['consider_prior'], prior_weight = cfg[cfg['optimizer']]['prior_weight'],n_startup_trials=cfg[cfg['optimizer']]['n_startup_trials'], n_ehvi_candidates=cfg[cfg['optimizer']]['n_ei_candidates'])
        if cfg['mode'] == 'single': 
            defined_pruner=optuna.pruners.HyperbandPruner()
        
    optuna.logging.set_verbosity(10)
    optuna.logging.enable_propagation()
    optuna.logging.enable_default_handler()
    formatted_string = f"postgresql://jesse_user:{_cfg_['POSTGRES_PASSWORD']}@{_cfg_['POSTGRES_HOST']}/optuna_db"
    port = 8082

    if not is_port_in_use(port):
        c = ['optuna-dashboard', formatted_string, '--port', str(port)]
        handle = subprocess.Popen(c, stdin=subprocess.PIPE, stderr=subprocess.PIPE, stdout=subprocess.PIPE, shell=False)
        url = f'http://127.0.0.1:8082/'  
        webbrowser.open_new_tab(url)  
        print("Optuna dashboard started.")
    else:
        print("Optuna dashboard is already running.")

    if continueExistingStudy:
        if cfg['mode'] == 'multi':
            study_name = f'{study_name}-multi'
            if cfg['dual_mode'] == 'maximize':
                study = JoblibStudy(study_name=study_name, directions=["maximize", "maximize"], sampler=sampler,
                            storage=storage, load_if_exists=True,pruner=defined_pruner)
            else:
                study = JoblibStudy(study_name=study_name, directions=["maximize", "minimize"], sampler=sampler,
                            storage=storage, load_if_exists=True,pruner=defined_pruner)

        else:
            study_name = f'{study_name}-single'
            study = JoblibStudy(study_name=study_name, direction="maximize", sampler=sampler,
                        storage=storage, load_if_exists=True,pruner=defined_pruner)
    else:
        if cfg['mode'] == 'multi':
            study_name = f'{study_name}-multi'
            try:
                optuna.study.delete_study(study_name=study_name, storage=storage)
            except Exception as e:
                print(e)
                pass
            if cfg['dual_mode'] == 'maximize':
                study = JoblibStudy(study_name=study_name, directions=["maximize", "maximize"], sampler=sampler,
                            storage=storage, load_if_exists=False,pruner=defined_pruner)
            else:
                study = JoblibStudy(study_name=study_name, directions=["maximize", "minimize"], sampler=sampler,
                            storage=storage, load_if_exists=False,pruner=defined_pruner)
        else:   
            study_name = f'{study_name}-single'
            try:
                optuna.study.delete_study(study_name=study_name, storage=storage)
            except Exception as e:
                print(e)
                pass
            study = JoblibStudy(study_name=study_name, direction="maximize", sampler=sampler,
                                        storage=storage, load_if_exists=False,pruner=defined_pruner)                            

    cfg['study_full_name'] = study_name
    cfg['start_time'] = jh.now_to_timestamp()
    cfg['n_trials_start'] = study.trials[-1].number if study.trials else 0
    cfg['main_pid'] = os.getpid()
    study.set_user_attr("strategy_name", cfg['strategy_name'])
    study.set_user_attr("exchange", cfg['route'][0]['exchange'])
    study.set_user_attr("symbol", cfg['route'][0]['symbol'])
    study.set_user_attr("timeframe", cfg['route'][0]['timeframe'])
    study.set_user_attr("timespan_train", cfg['timespan-train']['start_date'] + " -> " + cfg['timespan-train']['finish_date'])
    study.set_user_attr("timespan_testing", cfg['timespan-testing']['start_date'] + " -> " + cfg['timespan-testing']['finish_date'])

    write_dict_to_yaml(cfg)
    #df = (study.trials_dataframe())
    study.optimize(objective, n_jobs=cfg['n_jobs'], n_trials=cfg['n_trials'])
    if process_status() == 'started':
        sync_publish('progressbar', {
            'current': 100,
            'estimated_remaining_seconds': 0
        },'optuna')
        sync_publish('alert', {
            'message': f"Finished {cfg['n_trials']} iterations. Check your best DNA candidates, "
                       f"if you don't like any of them, try modifying your strategy.",
            'type': 'success'
        }, 'optuna')
        
        #analysis(study)
    
    
def analyze()-> None:
    validate_cwd()
    cfg = get_config()
    start_date = cfg['timespan-train']['start_date']
    finish_date = cfg['timespan-train']['finish_date']
    storage = f"postgresql://{cfg['postgres_username']}:{cfg['postgres_password']}@{cfg['postgres_host']}:{cfg['postgres_port']}/{cfg['postgres_db_name']}"
    study_name = f"{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{start_date}-{finish_date}-{cfg['mode']}"
    try:
        study = optuna.load_study(study_name=study_name, storage=storage)
    except: 
        print(Fore.RED + '\nNo Study Found\n' + Style.RESET_ALL)
        exit(1)
    analysis(study,True)
    
def analysis(study,save_time=False):
    validate_cwd()
    cfg = get_config()
    
    # train_path = f"./storage/jesse-optuna/training_metrics/{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['type']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{cfg['timespan-train']['start_date']}-{cfg['timespan-train']['finish_date']}-{cfg['optimizer']}-{len(cfg['route'])} Pair"
    # test_path = f"./storage/jesse-optuna/testing_metrics/{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['type']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{cfg['timespan-train']['start_date']}-{cfg['timespan-train']['finish_date']}-{cfg['optimizer']}-{len(cfg['route'])} Pair"
    
    # if not os.path.exists(train_path):
        # os.makedirs(train_path)
    # if not os.path.exists(test_path):
        # os.makedirs(test_path)
    
    # source_train_path = f"//root/Installs/jesse-optuna-master/temp/training_metrics/{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['type']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{cfg['timespan-train']['start_date']}-{cfg['timespan-train']['finish_date']}-{cfg['optimizer']}-{len(cfg['route'])} Pair"
    # source_test_path = f"//root/Installs/jesse-optuna-master/temp/testing_metrics/{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['type']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{cfg['timespan-train']['start_date']}-{cfg['timespan-train']['finish_date']}-{cfg['optimizer']}-{len(cfg['route'])} Pair"
    
    # files1 = os.listdir(source_train_path)
    # files2 = os.listdir(source_test_path) 
    
    # for fname in files1: 
        # shutil.copy2(os.path.join(source_train_path,fname),train_path)
        # os.remove(os.path.join(source_train_path,fname))
    # for fname in files2: 
        # shutil.copy2(os.path.join(source_test_path,fname),test_path)
        # os.remove(os.path.join(source_test_path,fname))
        
    start_date = cfg['timespan-train']['start_date']
    finish_date = cfg['timespan-train']['finish_date']
    study_name = f"{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{start_date}-{finish_date}"
    with open(f'./storage/jesse-optuna/results/{study_name}-results.txt', "w+") as f:
        f.write(f"Number of finished trials: {len(study.trials)}\n")
        trials = sorted(study.best_trials, key=lambda t: t.values)
        final_trials = []
        for trial in trials:
            duplicate = False
            for final in final_trials:
                if trial.params == final.params:
                    duplicate = True
                    break
            if duplicate == False:
                final_trials.append(trial) 
                
        for trial in final_trials:
            print(f"Trial #{trial.number} Values: { trial.values} {trial.params}")
        for trial in final_trials:
            f.write(
                f"Trial: {trial.number} Values: {trial.values} Params: {trial.params}\n")
    global done
    global update_1 
    global update_2
    global update_3
    global update_4
    global update_full 
    global final_update
    global update_filt
    final_update = False
    update_1 = False
    update_2 = False
    update_3 = False
    update_4 = False
    update_full = False
    update_filt = False
    done = False

    # study_summaries = optuna.study_get_all_study_summaries(storage=storage)
    # try:
    #Trial sorting assuming ['maximize']['maximize'] or ['maximize'] directions.
    update_filt = True
    
    trials = study.get_trials()
    candidate_trial_list = []
    average_testing_score_1 = 0
    average_testing_score_2 = 0
    param_list = []
    for trial in trials: 
        # reject pruned trials
        if trial.state != optuna.trial.TrialState.COMPLETE:
            continue
        # Don't save trials with existing parameter combinations.
        if trial.params in param_list:
            continue
        param_list.append(trial.params)
        if cfg['mode'] == 'multi':
            if cfg['dual_mode'] == "maximize":
                if trial.user_attrs['testing_score_1'] < trial.user_attrs['training_score_1'] and trial.user_attrs['testing_score_2'] < trial.user_attrs['training_score_2']:
                    continue
            else:
                if trial.user_attrs['testing_score_1'] < trial.user_attrs['training_score_1'] and trial.user_attrs['testing_score_2'] > trial.user_attrs['training_score_2']:
                    continue
        else:
            if trial.user_attrs['testing_score_1'] < trial.user_attrs['training_score_1']:
                continue 
        # check to see if testing draw-down is less than training draw_down         
        # if trial.user_attrs['testing-max_drawdown'] > trial.user_attrs['training-max_drawdown']:
            # continue 
            
        candidate_trial_list.append(trial)
        average_testing_score_1 += trial.user_attrs['testing_score_1']
        if cfg['mode'] == 'multi':
            average_testing_score_2 += trial.user_attrs['testing_score_2']
    average_testing_score_1 = average_testing_score_1/(len(candidate_trial_list))
    if cfg['mode'] == 'multi':
        average_testing_score_2 = average_testing_score_2/(len(candidate_trial_list))
    sorted_candidate_trial_list = []
    for trial in candidate_trial_list:
        if cfg['mode'] == 'multi':
            if cfg['dual_mode'] == 'maximize':
                if trial.user_attrs['testing_score_1'] < average_testing_score_1 and trial.user_attrs['testing_score_2'] < average_testing_score_2:
                    continue
            else:
                if trial.user_attrs['testing_score_1'] < average_testing_score_1 and trial.user_attrs['testing_score_2'] > average_testing_score_2:
                    continue
        else:
            if trial.user_attrs['testing_score_1'] < average_testing_score_1:
                continue 
        sorted_candidate_trial_list.append(trial)   
        
    dir1 = f"./storage/jesse-optuna/validation_metrics/{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{cfg['Interval_start']}-{cfg['Interval_end']}-{cfg['optimizer']}-{len(cfg['route'].items())} Pairs/"  
    sorted_candidate_trial_list_removal = []
    if os.path.exists(f'{dir1}removed_candidates.pkl') and save_time:
        with open(f'{dir1}removed_candidates.pkl', 'rb') as fp:
            sorted_candidate_trial_list_removal = pickle.load(fp)
    sorted_candidate_trial_list = [x for x in sorted_candidate_trial_list if x not in sorted_candidate_trial_list_removal]
    update_filt = False
        
    """
    Interval backtesting analysis
    """
    update_1 = True
    if not os.path.exists(dir1):
        os.mkdir(dir1)
    if os.path.exists(f'{dir1}temp'):
        try: 
            for file in os.listdir(f'{dir1}temp'):
                os.remove(f'{dir1}temp/{file}')
        except:
            pass
        os.rmdir(f'{dir1}temp')
    try: 
        for file in os.listdir(dir1):
            if file != f'removed_candidates.pkl':
                os.remove(f'{dir1}{file}')
    except:
        pass
        
    if not os.path.exists(f'{dir1}temp'):
        os.mkdir(f'{dir1}temp')
        
    Interval_start = cfg['Interval_start']
    Interval_end = cfg['Interval_end']
    
    for trial in study.best_trials:
        sorted_candidate_trial_list.append(trial)  
    sorted_candidate_trial_list = list(uniq(sorted(sorted_candidate_trial_list, reverse=True)))
    missing_files = False
    extra_files = False
    
    parallel = Parallel(n_jobs=cfg['n_jobs'],  backend='loky', verbose=0, max_nbytes=None)
    with parallel:
        results = parallel(delayed(metrics_func)(best_trial,missing_files,extra_files,cfg,Interval_start,Interval_end,dir1)
                    for best_trial in sorted_candidate_trial_list)
    delete_files = glob.glob(f'{dir1}/*-delete')
    if delete_files:
        for file in delete_files:
            file = file.split('/')[-1]
            file = file.split('-')[0]
            try:
                for file1 in sorted_candidate_trial_list:
                    if file1.number == file:
                        sorted_candidate_trial_list.remove(file1)
            except Exception as e:
                print(e)
            
    extra_files = results[1]
    missing_files = results[0]
    # delete all files less than 2KB because they are temporary files
    if extra_files and missing_files:
        for file in os.listdir(dir1):
            if file != f'removed_candidates.pkl':
                file = f"{dir1}{file}"
                if os.path.getsize(file) < (2 * 1024):
                    os.remove(file)
    try:
        os.rmdir(f'{dir1}temp')
    except:
        for filename in os.listdir(f'{dir1}temp/'):
            os.remove(filename)
        os.rmdir(f'{dir1}temp')
    
    update_1 = False
    """
    Full metrics backtesting 
    """
    update_full = True
    full_metrics_dict = {}
    if not os.path.exists(f'{dir1}temp'):
        os.mkdir(f'{dir1}temp')
    if not os.path.exists('./storage/charts/jesse-optuna'):
        os.mkdir('./storage/charts/jesse-optuna')
    if missing_files:
        with parallel:
            parallel(delayed(full_metrics_func)(dir1,best_trial, Interval_start, Interval_end, cfg)
                for best_trial in sorted_candidate_trial_list)
        delete_files = glob.glob(f'{dir1}/*-delete')
        if delete_files:
            for file in delete_files:
                file = file.split('/')[-1]
                file = file.split('-')[0]
                try:
                    for file1 in sorted_candidate_trial_list:
                        if file1.number == file:
                            sorted_candidate_trial_list.remove(file1)
                except Exception as e:
                    print(e)
        for filename in os.listdir(f"{dir1}temp/"):
            if filename.endswith(".pkl") and file != f'removed_candidates.pkl':
                with open(f'{dir1}temp/{filename}', 'rb') as f:
                    temp_trial_num = filename.split('.')[0]
                    full_metrics_dict[f"{temp_trial_num}"] = pickle.load(f)
                os.remove(f'{dir1}temp/{filename}')
        for file in os.listdir(dir1):
            if file != f'removed_candidates.pkl':
                file = f"{dir1}{file}"
                if os.path.getsize(file) < (100):
                    os.remove(file)
    os.rmdir(f'{dir1}temp')
    
    update_full = False
    """
    Filtering Results
    """
    update_filt = True
    
    if missing_files: 
        sorted_path =  f"{dir1}final-trial:*.csv"
        sorted_files = glob.glob(sorted_path)
        output_df = pd.DataFrame(columns=['trial_number','smart_sharpe_mean'])
        #smart_sharpe score weighted so last 1/3 of backtest accounts for 1/2 of mean_score 
        for filename in sorted_files:
            filename_str = os.path.basename(filename)
            filename_str = filename_str.split(':')[1]
            trial_num_str = str(filename_str.split('-')[0])
            temp_df = pd.read_csv(filename)
            # print(temp_df)
            mean_weighting = (math.floor(temp_df.shape[1]*0.60))
            new_df_weighted = temp_df.iloc[31,mean_weighting:temp_df.shape[1]]
            new_df = temp_df.iloc[31,1:temp_df.shape[1]]
            # print(filename)
            # print(new_df)
            # print(new_df_weighted)
            new_df = new_df.astype(float)
            new_df_weighted = new_df_weighted.astype(float)
            mean_value = new_df.mean(axis=0)
            mean_value_weighted = new_df_weighted.mean(axis=0)
            #slightly less than half need to be filtered 
            weighted_mean = (((mean_value*1)+(mean_value_weighted*1))/2)
            # print(f'mean_value: {mean_value}')
            # print(f'mean_value_weighted: {mean_value_weighted}')
            # print(f'weighted_mean: {weighted_mean}')
            output_df = output_df.append({'trial_number':trial_num_str,'smart_sharpe_mean':mean_value}, ignore_index=True)
            # exit(1)
        #quantile set to have twenty or less final results
        final_num = cfg["max_final_number_of_validation_results"]
        if (len(output_df['trial_number'].tolist())) > final_num:
            for i in range(1000,-1,-1):
                q_percent = (i * 0.001)
                smart_sharpe_mean_average = output_df.iloc[:,1].quantile(q_percent)*1 #output_df.iloc[:,1].median()*1
                temp_df1 = output_df[~(output_df['smart_sharpe_mean'] < smart_sharpe_mean_average)].reset_index()
                if len(temp_df1['trial_number'].tolist()) >= final_num: 
                    break
        else:
            smart_sharpe_mean_average = output_df.iloc[:,1].quantile(0.5)*1 #~ half if less than final_num
        if len(sorted_candidate_trial_list_removal) > final_num and len(sorted_candidate_trial_list_removal) > 0 or len(sorted_candidate_trial_list_removal) == 0:
            output_df = output_df[~(output_df['smart_sharpe_mean'] < smart_sharpe_mean_average)].reset_index()
        else:   
            output_df = output_df.reset_index()
        count = len(sorted_files)
        for f in os.listdir(dir1):
            try:
                if f != f'removed_candidates.pkl':
                    filename_str = f.split(':')[1]
                    trial_num_str = str(filename_str.split('-')[0])
                    filename = f"{dir1}{f}"
                    if any(str(elem) in [trial_num_str] for elem in output_df['trial_number'].tolist()):
                        continue 
                    else:
                        os.remove(filename)
            except IndexError:
                continue
                
        update_filt = False
        """
        Robust Parameter Testing
        """
        update_2 = True
          
        output_df = output_df.sort_values(by=['smart_sharpe_mean'],axis=0,ascending=[True],ignore_index=True)     
        output_df = output_df.iloc[:,1:]
        for item in sorted_candidate_trial_list:  
            found = False
            for elem in output_df['trial_number'].tolist():
                if float(item.number) == float(elem):
                    found = True
                    break
            if found == False and item not in sorted_candidate_trial_list_removal:
                sorted_candidate_trial_list_removal.append(item)
        sorted_candidate_trial_list = [x for x in sorted_candidate_trial_list if x not in sorted_candidate_trial_list_removal]
        if not save_time:
            with open(f'{dir1}removed_candidates.pkl', 'wb') as fp:
                pickle.dump(sorted_candidate_trial_list_removal, fp)
        # remove files not in sorted_candidate_trial_list
        for item in os.listdir(dir1):
            try:
                if item != f'removed_candidates.pkl':
                    found = False
                    for trial in sorted_candidate_trial_list:
                        num = trial.number 
                        file = item.split(':')[1]
                        if str(num) == str(file.split('-')[0]):
                            found = True
                            break
                    if found == False:
                        os.remove(f'{dir1}{item}')
            except IndexError:
                continue
        file_count = sum(1 for item in os.listdir(dir1) if isfile(join(dir1, item))) 
        # for trial in sorted_candidate_trial_list:
            # print(trial.number)
        with parallel:
            parallel(delayed(random_params_func)(dir1, output_df, full_metrics_dict,cfg,index)
                for index in range(0,len(sorted_candidate_trial_list)))
            
        update_2 = False
                    
        """
        Random Pairs Testing
        """ 
        update_4 = True
        
        with parallel:
            parallel(delayed(random_pairs_func)(dir1, output_df, cfg, index, sorted_candidate_trial_list)
                     for index in range(len(sorted_candidate_trial_list)))
        
        update_4 = False

        """
        Multiple Timeframe Testing and Final Formatting
        """
        
        update_3 = True
        
        from jesse.config import config as jesse_config
        
        jesse_config['env']['simulation']['skip'] = False

        with parallel:
            parallel(delayed(multiple_timeframes_func)(dir1, output_df, cfg, index, sorted_candidate_trial_list)
                     for index in range(len(sorted_candidate_trial_list)))
            
        jesse_config['env']['simulation']['skip'] = True
        
        update_3 = False 
    
    """
    QuantStat Reports
    """
    
    final_update = True  
    
    # moved to multiple_timeframes_func
    # parallel(delayed(quant_func)(dir1,best_trial,Interval_start,Interval_end,cfg,run_silently=False)
        # for best_trial in sorted_candidate_trial_list)   
        
    for chart in os.listdir('./storage/charts/jesse-optuna/'):
        try:
            os.remove(f'./storage/charts/jesse-optuna/{chart}')
        except:
            continue
    for chart in os.listdir('./storage/charts/'):
        if os.path.isfile(f'./storage/charts/{chart}'):
            try:
                os.remove(f'./storage/charts/{chart}')
            except:
                continue
    html_files = glob.glob(f"{dir1}*.html")
    excel_files = glob.glob(f"{dir1}*.xlsx")  
    try:
        assert len(html_files) == len(excel_files)
    except: 
        print(Fore.RED + Style.BRIGHT + "Missing QuantStat or Excel Files")
    
    importance = optuna.importance.get_param_importances(
        study, target= lambda t: t.values[0], evaluator=FanovaImportanceEvaluator()
    )
    
    try:
        if cfg['mode'] == 'multi': 
            fig1 = optuna.visualization.plot_pareto_front(study, targets=lambda t: (t.values[0], t.values[1]), target_names= [cfg['fitness-ratio1'], cfg['fitness-ratio2']])
        else: 
            fig1 = optuna.visualization.plot_pareto_front(study)
        fig1.show()
    except Exception as e:
        print(e)
        pass
    try:
        fig2 = optuna.visualization.plot_optimization_history(study, target= lambda t: t.values[0])
        fig2.show()
    except Exception as e:
        print(e)
        pass
    try: 
        fig5 = optuna.visualization.plot_contour(study, target= lambda t: t.values[0])
        fig5.show()
    except Exception as e:
        print(e)
        pass
    try:
        fig3 = optuna.visualization.plot_param_importances(study, target= lambda t: t.values[0])
        fig3.show()
    except Exception as e:
        print(e)
        pass
    try:
        fig4 = optuna.visualization.plot_slice(study, target= lambda t: t.values[0])
        fig4.show()
    except Exception as e:
        print(e)
        pass
    done = True
    t.join()
    
    print(f'Parameter Importance: {importance}\n')
    
    _text = Fore.MAGENTA + Style.BRIGHT + f"\n files are stored in C:/Python39/Algotrading{dir1}"
    print(_text)
    _text = Style.RESET_ALL + '\n '
    print(_text)

def uniq(lst):
    last = object()
    for item in lst:
        if item == last:
            continue
        yield item
        last = item
    
# def quant_func(dir1,best_trial,Interval_start,Interval_end,cfg,run_silently=False):  
    # params =  best_trial.params
    # trial_num = int(best_trial.number)
    # path_name = get_file_names_with_strings(dir1,[str(trial_num)])
    # path_name = path_name[0]
    # path_name = path_name.rsplit('-',11)[0]
    # get_quantreports = optuna_backtest_function_quant(Interval_start,Interval_end,params,cfg,run_silently=False,trial_num=trial_num,imported_string=path_name)
        
def get_file_names_with_strings(dir1,str_list):
    full_list = os.listdir(f'{dir1}')
    final_list = [nm for ps in str_list for nm in full_list if ps in nm]
    return final_list
    
def avrg(data,s):
    vals = [i[s] for i in data]
    return sum(vals)/len(vals)
    
def openpyxl_formating(df,base_filename,trial_num,cfg):
    # print(f'Trial Number: {trial_num}')
    wb = load_workbook(base_filename)
    ws = wb.active 
    ws.title = 'Optimization Result Analysis' 
    row_count = int(df.shape[0] + 1)  #100
    column_count = (df.count(axis='columns')[3]) + 1 #int(df.shape[1] + 1)
    none_columns = list(range(column_count, 42))
    # print(f"row_count: {row_count}")
    # print(f"column_count: {column_count}")
    bold_column1 = list(range(3,49))
    bold_column2 = list(range(52,98))
    random_rows = list(range(52,97))
    bold_header = list(range(1,column_count+1))
    special_bold_list = [50,51,2]
    data_columns = list(range(2,column_count+1))
    data_columns2 = [5,6,7,8,9,10,11,12,13,14]# list(range(5,13+1))
    data_columns3 = [21,22,23,24]
    # print(f"size of data_columns: {len(data_columns)}")
    win_trade_shade = [4]
    win_trade_shade2 = [53]
    loss_trade_shade = [5]
    loss_trade_shade2 = [54]
    win_rate_shade = [8]
    win_rate_shade2 = [57]
    pf_shade = [9]
    pf_shade2 = [58]
    avg_loss_shade = [19]
    avg_loss_shade2 = [68]
    omega_shade = [33]
    omega_shade2 = [82]
    losing_streak_shade = [40]
    losing_streak_shade2 = [89]
    shade_row1 = [7]
    shade_row2 = [56]
    legend_column = [9,11,12,14,15,17,18,20,21,23,24,26]
    second_top = [5,6,7,8,9,10,11,12,13,14]
    date_shade = [45]
    date_shade2 = [95,51]
    holding_shade = [23,24,25]
    holding_shade2 = [72,73,74]
    holding_shade3 = [73,74,75]
    reset_row = [23,25,37]
    fix_rows1 = [17,18]
    mid_column_id = list(range(51,96))
    mid_column_id2 = list(range(53,82))
    second_left_border = list(range(51,98))
    second_top_border_columns = list(range(2,15))  
    robust_list = list(range(102,143))
    # three point color scale based on time: 23,24,25
    big_shade = [16,17.18,20,21,22,26,27,28,29,30,31,32,34,35,36,38,39,41,42,43,44]
    big_shade2_2 = [56,53,54,57,58,68,82,89,65,66,67,69,70,71,75,76,77,78,79,80,81,83,84,85,87,88,90,91,92,93]
    big_shade2 = [65,66,67,68,69,70,71,75,76,77,78,79,80,81,82,83,84,85,87,88,89,90,91,92,93]
    big_shade4 = [66,67,68,69,70,71,72,76,77,78,79,80,81,82,83,84,85,86,88,89,90,91,92,93,94] 
    big_shade3_3 = [53,55,56,57,58,59,60,61,63,64,65,66,67,68,69,70,71,72,73,74,75]
    robust_border = list(range(53,82))
    chart_1_formatting = list(range(102,(101+(len(data_columns)))))
    chart_2_formatting = list(range(102, (102+(len(data_columns2)))))
    str_border = Border(right=Side(border_style='thin', color='FF000000'))
    top_border = Border(top=Side(border_style='thin', color='FF000000'))
    bottom_border = Border(bottom=Side(border_style='thin', color='FF000000'))
    left_border = Border(left=Side(border_style='thin', color='FF000000'))
    black_text = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,color='FF000000')
    black_bold_text = Font(name='Calibri', size=12, bold=False, italic=False, vertAlign=None, underline='none', strike=False,color='FF000000')
    special_bold = Font(name='Calibri', size=12, bold=True, italic=False, vertAlign=None, underline='none', strike=False,color='FF000000')
    red_text = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color="e32636")
    green_text = red_text = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color="32cd32")
    static_green_color = "008000"
    static_red_color = "FF0000"
    
    # find x and y deviation of pd.dataframe compared to excel cell
    for i in range(10):  
        row = 4 
        column = 3
        row2 = 6
        column2 = 5
        if ws.cell(row=row, column=column).value == df.iloc[row-i,column-1] and ws.cell(row=row2, column=column2).value == df.iloc[row2-i,column2-1]:
            x_deviation = i 
            break
    for i in range(10):
        row = 4
        column = 3 
        ro2 = 6 
        column2 = 5
        if ws.cell(row=row, column=column).value == df.iloc[row-(x_deviation),column-i] and ws.cell(row=row2, column=column2).value == df.iloc[row2-(x_deviation), column2-i]:
            y_deviation = i 
            break
    #temp
    y_deviation = 1
    
    temp_counter = 0
    for row in range(50):
        if not pd.isnull((df.iloc[(84+row-(x_deviation)),15])) and df.iloc[(84+row-(x_deviation)),15] is not str and temp_counter != cfg['robust_test_iteration_count']:
            temp_counter += 1 
        else:
            break
            
    chart_3_formatting = list(range(102, 102 + (temp_counter)))

    # print(f'X Dev: {x_deviation}')
    
    symbol_pairs = []
    for i in range(len(cfg['route'].items())):
        symbol_pairs.append(cfg['route'][i]['symbol'])
    symbol_pairs = ','.join(symbol_pairs)
    
    for row in range(1,200):
        for column in range(1,42):
            # print(f"current_row: {row}")
            # print(f"current_column: {column}")  
            if row == 1 and column in none_columns:
                ws.cell(row=row,column=column).value = ''
            if row == 48 and column in data_columns:
                if (column+1) in data_columns:
                    ws.cell(row=row,column=column).border = top_border    
            if row == 51 and column in second_top:
                ws.cell(row=row, column=column).border = top_border
            if row == 52 and (column == 20 or column in data_columns3):
                ws.cell(row=row, column=column).border = top_border
            if column == 20 and row in random_rows: 
                ws.cell(row=row,column=column).border = left_border  
            if (column == 16 or column == 4) and row == 50:
                ws.cell(row=row, column=column).font = special_bold
            if (column == 1 or column == 16) and row == 51:
                ws.cell(row=row, column=column).font = special_bold
            if column == 16 and row == 52:
                ws.cell(row=row, column=column).font = special_bold
            if (column == 16 and row == 83):
                ws.cell(row=row, column=column).font = special_bold
            if (column == 17 and row == 83):
                ws.cell(row=row, column=column).font = special_bold
            if (column == 18 and row == 83):    
                ws.cell(row=row, column=column).font = special_bold
            if (column == 4) and row in mid_column_id:
                ws.cell(row=row, column=column).font = black_bold_text
                ws.cell(row=row, column=column).border = str_border
            if (column == 4) and row in mid_column_id2:
                ws.cell(row=row, column=column).font = black_bold_text
            if row == 51 and column == 2:
                ws.cell(row=row,column=column).border = bottom_border
            if row == 52 and (column == 16 or column == 17):
                ws.cell(row=row, column=column).border = bottom_border
            if column == (len(data_columns)+1) and (row in bold_column1 and not row == 48) or (row == 3 and column == (len(data_columns)+1)):
                ws.cell(row=row,column=column).border = left_border   
            if column == 15 and row in second_left_border:
                ws.cell(row=row, column=column).border = left_border
            if column in second_top_border_columns and row == 98:
                ws.cell(row=row, column=column).border = top_border
            if row == 3 and column in data_columns:
                ws.cell(row=row, column=column).border = top_border
            if row == 99 and column == 6:
                ws.cell(row=row, column=column).value = "Reformatted Chart Data"
                ws.cell(row=row, column=column).font = special_bold
            if row == 2 and column in legend_column:   
                if row == 2 and column == 9:
                    ws.cell(row=row, column=column).value = "Most Recent/Frequent:"  
                if row == 2 and column == 11:
                    ws.cell(row=row,column=column).fill = PatternFill("solid", start_color = "030ffa", end_color = "030ffa", fill_type='solid')
                if row == 2 and column == 12:
                    ws.cell(row=row, column=column).value = "Least Recent/Frequent:"
                if row == 2 and column == 14:
                    ws.cell(row=row, column=column).fill = PatternFill("solid", start_color = "a1a7fc", end_color= "a1a7fc", fill_type='solid')
                if row == 2 and column == 15:
                    ws.cell(row=row, column=column).value = "Most Negative/Loss:"
                if row == 2 and column == 17:
                    ws.cell(row=row, column=column).fill = PatternFill("solid", start_color = "FF0000", end_color= "FF0000", fill_type='solid')
                if row == 2 and column == 18: 
                    ws.cell(row=row, column=column).value = "Least Negative/Loss:"
                if row == 2 and column == 20:
                    ws.cell(row=row, column=column).fill = PatternFill("solid", start_color = "8B0000", end_color="8B0000", fill_type='solid')
                if row == 2 and column == 21:
                    ws.cell(row=row, column=column).value = "Highest Positive/Gain:"
                if row == 2 and column == 23:
                    ws.cell(row=row, column=column).fill = PatternFill("solid", start_color="00cc00", end_color="00cc00", fill_type='solid')
                if row == 2 and column == 24:
                    ws.cell(row=row, column=column).value = "Lowest Positive/Gain:"
                if row == 2 and column == 26:
                    ws.cell(row=row, column=column).fill = PatternFill("solid", start_color = "024f02",end_color="024f02",fill_type='solid')
                continue 
            if row == 101: 
                if row == 101 and column == 3:
                    ws.cell(row=row,column=column).value = "Start Date"
                if row == 101 and column == 4:
                    ws.cell(row=row, column=column).value = "Net Profit Percentage"
                if row == 101 and column == 6:
                    ws.cell(row=row, column=column).value = "Timeframe" 
                if row == 101 and column == 7:
                    ws.cell(row=row, column=column).value = "Net Profit Percentage"
                if row == 101 and column == 9:
                    ws.cell(row=row, column=column).value = "Param % Deviation"
                if row == 101 and column == 10:
                    ws.cell(row=row, column=column).value = "Robust Net Profit % Per Iteration"
                if row == 101 and column == 12: 
                    ws.cell(row=row, column=column).value = "Symbol"
                if row == 101 and column == 13: 
                    ws.cell(row=row, column=column).value = "Net Profit % Per Symbol"
                
            # if row == 101 and column == 5:
                # ws.cell(row=row,column=column).value = 0
            # if row==102 and column ==5:
                # ws.cell(row=row,column=column).value = 1
            # if row == 101 and column == 6:
                # ws.cell(row=row,column=column).value = "Net Profit % per Iteration"
            # if row in robust_list and column == 6 or row in robust_list and column == 7:
                # if not ((ws.cell(row=(row-18), column= (column+10)).value)) == None :
                    # value = float(ws.cell(row=(row-18), column= (column+10)).value)
                    # if column == 7:
                        # row = row+1
                        # column = column-1
                        # ws.cell(row=row,column=column).value = value         
            if column == 3 and row in chart_1_formatting:
                # value = str(df.iloc[(45-(x_deviation)),((column-(y_deviation+1))+ -(102-row))])
                value = str(ws.cell(row=45,column=((column-1)+ -(102-row))).value)
                ws.cell(row=row, column=column).value = value 
            if column == 6 and row in chart_2_formatting:
                value = str(ws.cell(row=51,column=((column-1)+ -(102-row))).value)
                ws.cell(row=row, column=column).value = value 
            if column  == 12 and row in [102,103,104,105]:
                value = str(ws.cell(row=52,column=((column+9)+ -(102-row))).value)
                ws.cell(row=row, column=column).value = value 
            if column == 13 and row in [102,103,104,105]:
                value = (ws.cell(row=67,column=((column+8)+ -(102-row))).value)
                ws.cell(row=row, column=column).value = value 
            if column == 9 and row in chart_3_formatting:
                value = (ws.cell(row=(84 -(102-row)),column=16).value)
                ws.cell(row=row, column=column).value = value    
            if column == 10 and row in chart_3_formatting:
                value = (ws.cell(row=(84 -(102-row)),column=17).value)
                ws.cell(row=row, column=column).value = value
            if column == 7 and row in chart_2_formatting:
                value = (ws.cell(row=66,column=((column-2)+ -(102-row))).value)
                ws.cell(row=row, column=column).value = value 
            if column == 4 and row in chart_1_formatting:
                # value = float(df.iloc[(17-(x_deviation)),((column-(y_deviation+2))+ -(102-row))])
                # print(ws.cell(row=17,column=((column-2)+ - (102-row))).value)
                # print(base_filename)
                value = (ws.cell(row=17,column=((column-2)+ - (102-row))).value)
                value = float(value) if value != None else value
                ws.cell(row=row,column=column).value = value 
            if column == 1 and row == 48:
                ws.cell(row=row, column=column).value = "Tickers:"
            if column == 2 and row == 48:
                ws.cell(row=row, column=column).value = symbol_pairs
                
                
            # None Check for empty excel values 
            if ws.cell(row=row,column=column).value is None:
                continue
                
                
            if (row in bold_column1 or row in bold_column2) and column == 1:
                ws.cell(row=row, column=column).font = black_bold_text
                ws.cell(row=row, column=column).border = str_border
                
            if row == 1 and column in (bold_header):
                ws.cell(row=row, column=column).font = black_bold_text
                
            if (row in big_shade and column in data_columns) or (row in big_shade2 and column in data_columns2) or (row in big_shade4 and column in data_columns3):
                check = 0.0
                value = float(ws.cell(row=row, column=column).value)
                if column in data_columns and row in big_shade :
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values)
                else:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),4:14].astype(float).values)
                if column in data_columns3 and row in big_shade4:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),20:24].astype(float).values)
                    
                if value >= check:
                    color = green_shade_check(value,check,sorted_df1,trial_num)
                elif value < check:
                    color = red_shade_check(value,check,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
                
            if (row in win_trade_shade and column in data_columns) or (row in win_trade_shade2 and column in data_columns2) or (row == 54 and column in data_columns3):
                check = float(ws.cell(row=row+1, column = column).value)
                # float(df.iloc[(row-(x_deviation-1)),column-(y_deviation)])
                value = float(ws.cell(row=row, column=column).value)
                if column in data_columns and row in win_trade_shade:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values)
                else:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),4:14].astype(float).values)
                if column in data_columns3 and row == 54:   
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),20:24].astype(float).values)
                if value >= check:
                    color = green_shade_check(value,check,sorted_df1,trial_num)
                elif value < check:
                    color = red_shade_check(value,check,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color)
             
            if (row in loss_trade_shade and column in data_columns) or (row in loss_trade_shade2 and column in data_columns2) or (row == 55 and column in data_columns3):
                if column in data_columns and row in loss_trade_shade:
                    check = (np.median(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values))
                else:
                    check = (np.median(df.iloc[(row-(x_deviation)),4:14].astype(float).values))
                if column in data_columns3 and row == 55:
                    check = (np.median(df.iloc[(row-(x_deviation)),20:24].astype(float).values))
                value = -(float(ws.cell(row=row, column=column).value))
                if column in data_columns and (row in loss_trade_shade and column in data_columns):
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values)
                    sorted_df1 = [-x for x in sorted_df1]
                else:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),4:14].astype(float).values)
                    sorted_df1 = [-x for x in sorted_df1]
                if column in data_columns3 and row == 55:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),20:24].astype(float).values)
                    sorted_df1 = [-x for x in sorted_df1]
                if value >= 0:
                    color = green_shade_check(value,check,sorted_df1,trial_num)
                elif value < check:
                    color = red_shade_check(value,check,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color)
            
            
            if ((row in pf_shade or row in omega_shade or row in win_rate_shade) and column in data_columns) or ((row in pf_shade2 or row in omega_shade2 or row in win_rate_shade2) and column in data_columns2) or ((row == 58 or row == 59 or row == 83) and column in data_columns3):
                if row in win_rate_shade or row in win_rate_shade2 or row == 58:
                    check = 0.5
                else:
                    check = 1.0
                value = float(ws.cell(row=row, column=column).value)
                if column in data_columns and (row in pf_shade or row in omega_shade or row in win_rate_shade):
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values)
                else:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),4:14].astype(float).values)
                if column  in data_columns3 and (row == 58 or row == 59 or row == 83):
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),20:24].astype(float).values)
                if value >= check:
                    color = green_shade_check(value,check,sorted_df1,trial_num)
                elif value < check:
                    color = red_shade_check(value,check,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
                
                        
            if ((row in avg_loss_shade or row in losing_streak_shade) and column in data_columns) or ((row in avg_loss_shade2 or row in losing_streak_shade2) and column in data_columns2) or ((row == 90 or row == 69) and column in data_columns3):
                check = math.inf
                value = -(float(ws.cell(row=row, column=column).value))
                if column in data_columns and (row in avg_loss_shade or row in losing_streak_shade):
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values)
                    sorted_df1 = [-x for x in sorted_df1]
                else:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),4:14].astype(float).values)
                    sorted_df1 = [-x for x in sorted_df1]
                if column in data_columns3 and (row == 90 or row == 69):
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),20:24].astype(float).values)
                    sorted_df1 = [-x for x in sorted_df1]
                if value >= check:
                    color = green_shade_check(value,check,sorted_df1,trial_num)
                elif value < check:
                    color = red_shade_check(value,check,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
                    
                
            if (row in shade_row1 and column in data_columns) or (row in shade_row2 and column in data_columns2) or (row == 57 and column in data_columns3):
                check = float(ws.cell(row-1,column=column).value)
                # float(df.iloc[(row-(x_deviation+1)),column-(y_deviation)])
                value = float(ws.cell(row=row, column=column).value)
                if column in data_columns and (row in shade_row1 and column in data_columns):
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values)
                else:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),4:14].astype(float).values)
                if column in data_columns3 and row == 57:
                    sorted_df1 = list(df.iloc[(row-(x_deviation)),20:24].astype(float).values)
                if value >= check:
                    color = green_shade_check(value,check,sorted_df1,trial_num)
                elif value < check:
                    color = red_shade_check(value,check,sorted_df1,trial_num)
    
                # three color scale 
                # red, gray, green = [ spectra.html(x).to("lab") for x in ("red", "#CCC", "green") ]
                # polylinear_scale = spectra.scale([ red, gray, green ])
                # swatches(polylinear_scale.range(9))

                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color)
            
            if row in reset_row and column in data_columns:
                ws.cell(row=row, column=column).font = black_text
                
            if (row in date_shade and column in data_columns) or (row in date_shade2 and column in data_columns2):    
                if column in data_columns2 and row in date_shade2:
                    num = len(data_columns2)
                    color = date_shade_reversed(column,num)
                else:
                    num = len(data_columns)
                    color = date_shade_(column,num)
                ws.cell(row=row, column=column).font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
            
            if row in holding_shade and column in data_columns:
                value = float(ws.cell(row=row, column=column).value)
                sorted_df1 = list(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values)
                color = holding_time_shade(value,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
            
            if row in holding_shade2 and column in data_columns2:
                value = float(ws.cell(row=row, column=column).value)
                sorted_df1 = list(df.iloc[(row-(x_deviation)),4:14].astype(float).values)
                color = holding_time_shade(value,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
          
            if row in holding_shade3 and column in data_columns3:
                value = float(ws.cell(row=row, column=column).value)
                sorted_df1 = list(df.iloc[(row-(x_deviation)),20:24].astype(float).values)
                color = holding_time_shade(value,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
                
            if column == 1 and row in (special_bold_list):
                ws.cell(row=row, column=column).font = special_bold
            
            if column == 20 and row in [50,51]:
                ws.cell(row=row, column=column).font = special_bold 
                
            if row in fix_rows1 and column in data_columns:
                check = 0.0
                value = float(ws.cell(row=row, column=column).value)
                sorted_df1 = list(df.iloc[(row-(x_deviation)),1:column_count].astype(float).values)
                # print(f'max value {max_value}')
                # print(f'min value {min_value}')
                # print(f'value {value}')
                if value >= check:
                    color = green_shade_check(value,check,sorted_df1,trial_num)
                elif value < check:
                    color = red_shade_check(value,check,sorted_df1,trial_num)
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
                
            if row in big_shade2_2 and column == 2:
                check = 0.0
                if row == 57:
                    check = 0.5
                if row == 58 or row == 82:
                    check = 1
                value = float(ws.cell(row=row, column=column).value)
                if value >= check:
                    color = static_green_color
                elif value < check:
                    color = static_red_color
                if row == 89 and column == 2:
                    color = static_red_color
                if row == 68:
                    color = static_red_color
                if row == 56 and column == 2:
                    if float(df.iloc[(row-(x_deviation-1)),column-(y_deviation)]) < value:
                        color = static_green_color
                    else:
                        color = static_red_color
                if row == 54:
                    color = static_red_color
                if row == 53:
                    # print(f'row53: {value} - {float(df.iloc[(row-(x_deviation-1)),column-(y_deviation)])} - {trial_num} - {y_deviation}')
                    if float(df.iloc[(row-(x_deviation-1)),column-(y_deviation)]) > value:
                        color = static_red_color
                    else:
                        color = static_green_color

                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color) 
            if row in big_shade3_3 and column == 17:
                check = 0.0
                value = float(ws.cell(row=row,column=column).value)
                if row == 55:
                    check = 0.5
                if row == 67 or row == 56:
                    check = 1 
                if value >= check:
                    color = static_green_color
                elif value < check:
                    color = static_red_color
                if row == 53:
                    if float(df.iloc[(row-(x_deviation-2)),column-(y_deviation+3)]) < value:
                        color = static_green_color
                    else:
                        color = static_red_color
                if row == 61:
                    color = static_red_color
                if row == 63:
                    color = static_red_color
                if row == 72:
                    color = static_red_color
                ws.cell(row=row, column=column).font =  Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=color)     
            if row in robust_border and column == 17:
                ws.cell(row=row, column=column).border = left_border
            if row in robust_border and column == 16:
                ws.cell(row=row, column=column).font = black_bold_text
                
           # if 
                # cell_value = ws.cell(row=row,column=col).value
                # if cell_value is not isinstance(cell_value,str) and float(cell_value) < 0:
                
                    # ws.cell(row=row, column=col).font = red_text
                # else:
                    # ws.cell(row=row, column=col).font = black_text
                    
    letter = get_column_letter(column_count+4)
    letter2 = get_column_letter(column_count+1)
    letter3 = get_column_letter(column_count+13)
    
    series_name = extract_specific_part(base_filename)
    
    chart1 = LineChart()
    chart1.height=10
    chart1.width=20
    chart1.title = "Net Profit % over Backtested Intervals"
    chart1.style = 3 #8
    chart1.y_axis.title = "Net Profit Percent"
    chart1.x_axis.title = "Interval Start Dates"
    row_end = 100+ (len(data_columns))
    chart1_data = Reference(ws, min_col = 4, min_row = 102, max_col = 4, max_row = row_end)
    dates = Reference(ws, min_col = 3, min_row = 102, max_col = 3, max_row = row_end)
    set_chart_title_size(chart1,size=1400)
    chart1.add_data(chart1_data, titles_from_data=False)
    chart1.set_categories(dates)
    s1 = chart1.series[0]
    s1.marker.symbol = "diamond"
    # if len(chart1.series) > 0:
        # chart1.series[0].graphicalProperties.line.solidFill = "00AAAA"  
        # series1_title_str_ref = StrRef()
        # series1_title_str_ref.f = series_name
        
        # series1_title = SeriesLabel()
        # series1_title.strRef = series1_title_str_ref
        # chart1.series[0].title = series1_title
    # chart1.dataLabels = DataLabelList()
    # chart1.dataLabels.showVal = True
    
    ws.add_chart(chart1, f"{letter2}4")
    
    series_name2 = extract_desired_part_no_time(base_filename)
    
    chart2 = BarChart()
    chart2.type = "bar"
    chart2.grouping = "clustered"
    chart2.height=10
    chart2.width=20
    chart2.title = "Net Profit % over Backtested Timeframes"
    chart2.style = 15 #5
    chart2.y_axis.title = "Net Profit Percent"
    chart2.x_axis.title = "Timeframe"
    row_end = 111 #101+ (len(data_columns2))
    chart2_data = Reference(ws, min_col = 7, min_row = 102, max_col = 7, max_row = row_end)
    dates = Reference(ws,min_col=6, min_row=102,max_col=6,max_row=row_end)
    set_chart_title_size(chart2,size=1400)
    chart2.add_data(chart2_data, titles_from_data=False)
    chart2.set_categories(dates)
    # if len(chart2.series) > 0:
        # chart2.series[0].graphicalProperties.line.solidFill = "00AAAA"  
        # series2_title = SeriesLabel()
        # series2_title.strRef = series_name2
        # chart2.series[0].title = series2_title
    # s2 = chart2.series[0]
    # s2.marker.symbol = "diamond"
    # chart2.dataLabels = DataLabelList()
    # chart2.dataLabels.showVal = True
    ws.add_chart(chart2, f"{letter2}23")
    
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.grouping = "clustered"
    chart3.height=10
    chart3.width=20
    chart3.title = "Net Profit % with Random Parameter Deviations"
    chart3.style = 13 #5
    chart3.y_axis.title = "Net Profit Percent"
    chart3.x_axis.title = "Param % Deviation"
    row_end = 101 + temp_counter
    chart3_data = Reference(ws, min_col = 10, min_row = 102, max_col = 10, max_row = row_end)
    dates = Reference(ws,min_col=9, min_row=102,max_col=9,max_row=row_end)
    set_chart_title_size(chart3,size=1400)
    chart3.add_data(chart3_data, titles_from_data=False)
    chart3.set_categories(dates)
    # if len(chart3.series) > 0:
        # chart3.series[0].graphicalProperties.line.solidFill = "00AAAA"  
        # series3_title = SeriesLabel()
        # series3_title.strRef = series_name
        # chart3.series[0].title = series3_title
    # s2 = chart2.series[0]
    # s2.marker.symbol = "diamond"
    # chart2.dataLabels = DataLabelList()
    # chart2.dataLabels.showVal = True
    ws.add_chart(chart3, f"{letter3}4")
    
    series_name3 = extract_specific_part_simple(base_filename)
    
    chart4 = BarChart()
    chart4.type = "bar"
    chart4.grouping = "clustered"
    chart4.height=10
    chart4.width=20
    chart4.title = "Net Profit % with Random Symbols"
    chart4.style = 10 #5
    chart4.y_axis.title = "Net Profit Percent"
    chart4.x_axis.title = "Symbol"
    row_end = 101 + 4
    chart4_data = Reference(ws, min_col = 13, min_row = 102, max_col = 13, max_row = row_end)
    dates = Reference(ws,min_col=12, min_row=102,max_col=12,max_row=row_end)
    set_chart_title_size(chart4,size=1400)
    chart4.add_data(chart4_data, titles_from_data=False)
    chart4.set_categories(dates)
    # if len(chart4.series) > 0:
        # chart4.series[0].graphicalProperties.line.solidFill = "00AAAA"  
        # series4_title = SeriesLabel()
        # series4_title.strRef = series_name3
        # chart4.series[0].title = series4_title
    # s2 = chart2.series[0]
    # s2.marker.symbol = "diamond"
    # chart2.dataLabels = DataLabelList()
    # chart2.dataLabels.showVal = True
    ws.add_chart(chart4, f"{letter3}23")
    try:
        png_loc = f'./storage/charts/jesse-optuna/{trial_num}-chart.png'
        my_png = Image(png_loc)
        ws.add_image(my_png, f'{letter}50')
    except Exception as e:
        print(f'unable to create a chart for {trial_num}')
        print(e)
        pass
    os.remove(base_filename)
    wb.save(base_filename)                    

def set_chart_title_size(chart, size=1400):
    paraprops = ParagraphProperties()
    paraprops.defRPr = CharacterProperties(sz=size)

    for para in chart.title.tx.rich.paragraphs:
        para.pPr=paraprops 
        
def extract_desired_part_no_time(filename):
    try:
        parts = filename.split('final-trial')
        if len(parts) > 1:
            interval_part = parts[1].split('-interval')
            if len(interval_part) > 0:
                subparts = interval_part[0].split('-')
                desired_part = '-'.join(subparts[2:-4]) + '-' + '-'.join(subparts[-3:])
                return desired_part
    except:
        return "Series1"

def extract_specific_part_simple(filename):
    try:
        parts = filename.split('final-trial')
        if len(parts) > 1:
            interval_part = parts[1].split('-interval')
            if len(interval_part) > 0:
                subparts = interval_part[0].split('-')
                desired_part = subparts[2] + '-' + '-'.join(subparts[-3:])
                return desired_part

    except:
        return "Series1"
    
def holding_time_shade(value,sorted_df,trial_num):
    try:
        sorted_df.sort()
        max_value = max(sorted_df)
        min_value = min(sorted_df)         
        sorted_df_index = list(np.arange(len(sorted_df)))
        sorted_df_dict = dict(zip(sorted_df_index,sorted_df))
        sorted_df_dict = dict(sorted(sorted_df_dict.items(), key = lambda kv: kv[1]))
        num = len(sorted_df)
        if num > 1:
            color_range = (spectra.range(["#030ffa","#a1a7fc"],num))
        else:
            return "000000"
        order_key = ([k for k, v in sorted_df_dict.items() if v == value])        
        order_key = order_key[0]
        color = color_range[order_key]
        color = str(color.hexcode)
        color = color.split('#')[1]
        return color       
    except Exception as e:
        print(f'holding_time_shade: {sorted_df} from {trial_num}')
        print(e)
        pass
 
def green_shade_check(value,check,sorted_df,trial_num):
    try:
        sorted_df = [num for num in sorted_df if num >= check]
        sorted_df.sort()
        max_value = max(sorted_df)
        min_value = min(sorted_df)         
        sorted_df_index = list(np.arange(len(sorted_df)))
        sorted_df_dict = dict(zip(sorted_df_index,sorted_df))
        sorted_df_dict = dict(sorted(sorted_df_dict.items(), key = lambda kv: kv[1]))
        num = len(sorted_df)
        if num > 1:
            color_range = (spectra.range(["#024f02","#00CC00"],num))
        else:
            return "00CC00"
        order_key = ([k for k, v in sorted_df_dict.items() if v == value])        
        order_key = order_key[0]
        color = color_range[order_key]
        color = str(color.hexcode)
        color = color.split('#')[1]
        return color       
    except Exception as e:
        print(f'green_shade_check failure: {sorted_df} from {trial_num} - check: {check} - value: {value}')
        print(e)
        pass
        
    
def date_shade_(column,num):
    column = column - 1
    color_range = (spectra.range(["#a1a7fc","#030ffa"],num))
    color = color_range[column]
    color = str(color.hexcode)
    color = color.split('#')[1]
    return color

def date_shade_reversed(column,num):
    column = column - 5
    color_range = (spectra.range(["#030ffa","#a1a7fc"],num))
    color = color_range[column]
    color = str(color.hexcode)
    color = color.split('#')[1]
    return color
    
def red_shade_check(value,check,sorted_df,trial_num):
    try:
        sorted_df = [num for num in sorted_df if num < check]
        sorted_df.sort()
        max_value = max(sorted_df)
        min_value = min(sorted_df)         
        sorted_df_index = list(np.arange(len(sorted_df)))
        sorted_df_dict = dict(zip(sorted_df_index,sorted_df))
        sorted_df_dict = dict(sorted(sorted_df_dict.items(), key = lambda kv: kv[1]))
        num = len(sorted_df)
        if num > 1:
            color_range = (spectra.range(["#FF0000","#8B0000"],num))  
        else:   
            return "FF0000"
        order_key = ([k for k, v in sorted_df_dict.items() if v == value])        
        order_key = order_key[0]
        color = color_range[order_key]
        color = str(color.hexcode)
        color = color.split('#')[1]        
        return color   
    except Exception as e:
        print(f'red_shade_check : {sorted_df} from {trial_num}')
        print(e)
        pass
        
    
    
def fixed_red_colorshade(value,min_value,max_value):
    col_list = ["FF0000","F40000","EA0000","DF0000","D50000","CA0000","C00000", "B50000","AB0000", "A00000", "960000", "8B0000"]
    col = '800080'
    for i in range(len(col_list)):
        diff_ = ((max_value-min_value)/(12-i))
        if value == min_value:
            col = col_list[0]
            break
        elif value == max_value:
            col = col_list[-1]
            break
        elif value < (((max_value-min_value)/(12-i))+min_value):
            col = col_list[i]
            break
    return col 

def fixed_green_colorshade(value,min_value,max_value):
    #list accidently reversed 
    col_list = ["003300","004100","004F00","005D00","006B00","007900","008600","009400","00A200","00B000","00BE00","00CC00"]
    col = '800080'
    for i in range(len(col_list)):
        diff = ((max_value-min_value)/(12-i))
        if value == min_value:
            col = col_list[0]
            break
        elif value == max_value:
            col = col_list[-1]
            break
        elif value < (((max_value - min_value) / (12-i))+min_value):
            col = col_list[i]
            break
    return col 

def Convert(tup, di):
    for a, b in tup:
        di.setdefault(a, []).append(b)
    return di
    
def full_metrics_func(dir1, best_trial, Interval_start, Interval_end, cfg):
    try:
        full_metrics = optuna_backtest_function(Interval_start,Interval_end,best_trial.params,cfg,run_silently=True,trial_num=(int(best_trial.number)),optimizing=True)
        full_metrics = {key : round(full_metrics[key],3) for key in full_metrics}
        with open(f'{dir1}temp/{best_trial.number}.pkl', 'wb') as f:
            pickle.dump(full_metrics, f)
    except Exception as e:
        print(e)
        with open(f'{dir1}/{best_trial.number}-delete', 'w') as file:
            pass 
        pass
   
def multiple_timeframes_func(dir1, output_df, cfg, index, sorted_candidate_trial_list):
    pair_start = cfg['Interval_start']
    pair_end = cfg['Interval_end']
    for f in os.listdir(dir1):
        try:
            if f != f'removed_candidates.pkl':
                filename_str = f.split(':')[1]
                trial_num_str = str(filename_str.split('-')[0])
                if trial_num_str == str(output_df.iloc[index,0]):
                    filename = f"{dir1}{f}"    
                    break
        except IndexError:
            continue
    for trial in sorted_candidate_trial_list:
        if str(trial.number) == str(trial_num_str):
            params = trial.params
    index = output_df[output_df.iloc[:,0].str.contains(trial_num_str)].index.values
    try:
        _sharpe_mean = str(round(float(output_df.iloc[index,1].values),2))
    except:
        _sharpe_mean = str(round(float(output_df.iloc[index[0],1]),2))
    file_split = filename.split('/')[-1]
    new_filename = f"{dir1}SERENITY:{(_sharpe_mean)}-{file_split}"
    new_dataframe = pd.read_pickle(f'{filename}') 
    
    timeframe_list = ['3m','5m','15m','30m','45m','1h','2h','3h','4h','6h']
    timeframe_name = ['3 minutes', '5 minutes', '15 minutes', '30 minutes', '45 minutes', '1 hour', '2 hours', '3 hours', '4 hours', '6 hours']
    temp_df_list = []
    for index, timeframe in enumerate(timeframe_list):
        try: 
            temp = {}
            old_timeframe = cfg['route'][0]['timeframe']
            cfg['route'][0]['timeframe'] = timeframe
            temp = optuna_backtest_function(pair_start,pair_end,params,cfg,optimizing=True)
            temp['index'] = index+1
            temp = {key : round(temp[key],3) for key in temp}
            temp['timeframe'] = timeframe_name[index]
            cfg['route'][0]['timeframe'] = old_timeframe
            temp_df = pd.DataFrame.from_dict(temp.items(), orient='columns')
        except Exception as e:
            print(e)
            temp_df = temp_df_list[index-1]
        temp_df_list.append(temp_df)
    t_df = pd.concat(temp_df_list,ignore_index=True,axis=1)
    drop_idx = list(range(2,t_df.shape[1],2))
    drop_cols =  [j for i,j in enumerate(t_df.columns) if i in drop_idx]
    t_df = t_df.drop(drop_cols,axis=1) 
    os.remove(f'{filename}')
    for i in range(t_df.shape[0]):
        if i == 0:
            new_dataframe.iat[48,3] = f"Backtest Metrics from 3 Minute Timeframe to 6 hours from {pair_start} to {pair_end}"
        for j in range(t_df.shape[1]):
            new_dataframe.iat[49,(j+3)] = t_df.iat[43,(j)]
            new_dataframe.iat[50+i,(3+j)] = t_df.iat[(0+i),(j)]
    
    old_csv = new_filename.rpartition('-')[0]
    new_filename = new_filename.partition(".csv")[0]
    new_filename = f"{new_filename}.xlsx"
    
    new_dataframe.to_excel(new_filename, engine='openpyxl', header=True, index=False, encoding='utf-8')
    
    path_name = get_file_names_with_strings(dir1,[trial_num_str])
    path_name = path_name[0]
    path_name = path_name.rsplit('-',11)[0]
    Interval_start = cfg['Interval_start']
    Interval_end = cfg['Interval_end']
    try: 
        full_metrics_ = optuna_backtest_function_quant(Interval_start,Interval_end,params,cfg,additional_files=True,generate_charts=True,generate_equity_curve=True,run_silently=False,trial_num=int(trial_num_str),imported_string=path_name,full_path_name = dir1,full_name=path_name)
        full_metrics = full_metrics_['metrics']
        chart = full_metrics_['charts']
    # equity_curve = full_metrics_['equity_curve']
    # TODO: Create a saved graph form the equity curve data
    # path1 = f'./storage/charts/jesse-optuna/{best_trial.number}-equity_curve.png'
        path2 = f'./storage/charts/jesse-optuna/{trial_num_str}-chart.png'
        os.rename(chart,path2)
    except Exception as e: 
        print(f'unable to create quantstat data for {trial_num_str}')
        print(e)
        pass
    
    openpyxl_formating(new_dataframe,new_filename,trial_num_str,cfg)

    try:
        os.remove(old_csv)
    except Exception as e:
        print("could not remove old csv file")
        print(e)
        pass

def extract_specific_part(filename):
    try: 
        parts = filename.split('final-trial')
        if len(parts) > 1:
            interval_part = parts[1].split('-interval')
            if len(interval_part) > 0:
                desired_part = interval_part[0].strip('-')
                return desired_part
    except:
        return "Series1"
            
def random_params_func(dir1,output_df,full_metrics_dict,cfg,index):
    robust_length = cfg['robust_test_iteration_count']
    robust_min = 100*(cfg['robust_test_min'])
    robust_max = 100*(cfg['robust_test_max'])
    Interval_start = cfg['Interval_start']
    Interval_end = cfg['Interval_end']
    for f in os.listdir(dir1):
        filename_str = f.split(':')[1]
        trial_num_str = str(filename_str.split('-')[0])
        if trial_num_str == str(output_df.iloc[index,0]):
            filename = f"{dir1}{f}"    
            break
    index = output_df[output_df.iloc[:,0].str.contains(trial_num_str)].index.values
    try:
        _sharpe_mean = str(round(float(output_df.iloc[index,1].values),2))
    except:
        _sharpe_mean = str(round(float(output_df.iloc[index[0],1]),2))
    file_split = filename.split('/')[-1]
    new_filename = f"{dir1}SERENITY:{(_sharpe_mean)}-{file_split}"
    temp_df = pd.read_csv(filename, header=None)
    params = temp_df.iloc[-1,1]
    _full_metrics = full_metrics_dict[trial_num_str]
    _full_metrics['start_date'] = Interval_start
    _full_metrics['end_date'] = Interval_end
    _full_metrics['index'] = 1 
    _full_metrics['parameters'] = params
    _full_metrics = pd.DataFrame.from_dict(_full_metrics.items(), orient='columns')
    for i in range(2):
        temp_df = temp_df.append(pd.Series(), ignore_index = True)
    new_roll = pd.Series([math.nan])
    new_roll.loc[0] = f"A Series of Backtests with a Interval of {cfg['validation_interval']} days from {Interval_start} to {Interval_end}"
    temp_df = pd.concat([new_roll,temp_df], axis=0, ignore_index=True)
    temp_df.loc[48,0] = f"Full Backtest Metrics From"
    temp_df.loc[49,0] = f"{Interval_start} to {Interval_end}"
    new_dataframe = pd.concat([temp_df, _full_metrics], axis=0, ignore_index = True)
    if not os.path.isfile(new_filename):
        os.rename(filename,new_filename)
    params = ast.literal_eval(params)
    list_rand_average = []
    random_params = {}
    for index in range(robust_length):
        temp_list = []
        temp_list = list(params.values())
        _temp_list_key = params.keys()
        temp_list_rand = []
        for i in range(len(temp_list)):
            n = random.randint(robust_min,robust_max)
            if pr.prob(0.5,num=1) == True:
                n = 1 - (n/10000)
            else:
                n = 1 + (n/10000)
            temp_list_rand.append(n)
        for i in range(len(temp_list)):
            if temp_list_rand[i] < 1:
                list_rand_average.append((abs(temp_list_rand[i]-1)+1))
            elif temp_list_rand[i] > 1:
                list_rand_average.append(temp_list_rand[i])
            if isinstance(temp_list[i],int):
                n = temp_list_rand[i]
                if temp_list[i] <= 10:
                    if (n-1) < 0:
                        if pr.prob((1-n),num=1):
                            if pr.prob(0.5,num=1):
                                temp_list[i] = temp_list[i] + 1
                            else:
                                temp_list[i] = temp_list[i] - 1
                    elif(n-1) >= 0:
                        if pr.prob((n-1),num=1):
                            if pr.prob(0.5,num=1):
                                temp_list[i] = temp_list[i] + 1
                            else: 
                                temp_list[i] = temp_list[i] - 1
                else:
                    temp_list[i] = round(temp_list[i] * n)
            elif isinstance(temp_list[i],float):
                n = temp_list_rand[i]
                temp_list[i] = round((temp_list[i] * n),2)
            elif isinstance(temp_list[i],bool): 
                if temp_list[i] == False:
                    n = 1 - temp_list_rand[i] 
                    temp_list[i] = pr.prob(n,num=1)
                elif temp_list[i] == True:
                    n = temp_list_rand[i] 
                    temp_list[i] = pr.prob(n,num=1)
            else:
                print("param type error")
                exit(1)
        final_temp_list = dict(zip(_temp_list_key,temp_list))
        random_params[index] = final_temp_list
    # Get the % deviation avg per iteration 
    rand_dev = list_rand_average 
    final_rand_dev = []
    temp_counter = 0 
    div = len(list(params.values()))
    for i in range(1,len(rand_dev)+1):
        temp_counter += 1 
        if temp_counter % div == 0:
            temp_2 = ((sum(rand_dev[i-div:i])/div)-1)*100
            final_rand_dev.append(temp_2)
            
    list_rand_average = (sum(list_rand_average)/len(list_rand_average))-1
    #assuming a perfrect distribution of negative and postiive deviations. The true deviation is twice as large.
    list_rand_average = list_rand_average*100
    list_rand_average = round(list_rand_average,3)
    final_rand_dev = [round(item,3) for item in final_rand_dev]
    random_metrics_iter = {}
    robust_start =  Interval_start
    robust_end = Interval_end
    output_dict = {}
    for i in range(robust_length):
        robust_params = random_params[i]
        try:
            full_metrics_ = optuna_backtest_function(robust_start,robust_end,robust_params,cfg, optimizing=True)
            full_metrics_['random_params'] = robust_params
            random_metrics_iter[i] = full_metrics_
        except: 
            continue

    try:
        finishing_balance_list = [random_metrics_iter[i]['finishing_balance'] for i in random_metrics_iter]
        finishing_balance_mean = round(sum(finishing_balance_list)/ len(finishing_balance_list),3)  
    except:
        finishing_balance_mean = None
    output_dict['finishing_balance_mean'] = finishing_balance_mean
    try: 
        total_list = [random_metrics_iter[i]['total'] for i in random_metrics_iter]
        total_mean = round(sum(total_list)/ len(total_list),3)
    except:
        total_mean = None
    output_dict['total_mean'] = total_mean
    try:
        win_rate_list = [random_metrics_iter[i]['win_rate'] for i in random_metrics_iter]
        win_rate_mean = round(sum(win_rate_list)/ len(win_rate_list),3)
    except:
        win_rate_mean = None
    output_dict['win_rate_mean'] = win_rate_mean
    try:
        profit_factor_list = [random_metrics_iter[i]['profit_factor'] for i in random_metrics_iter]
        profit_factor_mean = round((sum(profit_factor_list)/ len(profit_factor_list)),3)
    except:
        profit_factor_mean = None
    output_dict['profit_factor_mean'] = profit_factor_mean
    try:
        ratio_avg_win_loss_list = [random_metrics_iter[i]['ratio_avg_win_loss'] for i in random_metrics_iter]
        ratio_avg_win_loss_mean = round(sum(ratio_avg_win_loss_list)/ len(ratio_avg_win_loss_list),3)
    except:
        ratio_avg_win_loss_mean = None
    output_dict['ratio_avg_win_loss_mean'] = ratio_avg_win_loss_mean
    try: 
        net_profit_list = [random_metrics_iter[i]['net_profit'] for i in random_metrics_iter]
        net_profit_mean = round(sum(net_profit_list)/ len(net_profit_list),3)
    except:
        net_profit_mean = None
    output_dict['net_profit_mean'] = net_profit_mean
    try: 
        net_profit_percentage_list = [random_metrics_iter[i]['net_profit_percentage'] for i in random_metrics_iter]
        net_profit_percentage_mean = round(sum(net_profit_percentage_list)/ len(net_profit_percentage_list),3)
    except:
        net_profit_percentage_mean = None
    output_dict['net_profit_percentage_mean'] = net_profit_percentage_mean
    try: 
        average_win_mean_list = [random_metrics_iter[i]['average_win'] for i in random_metrics_iter]
        average_win_mean = round(sum(average_win_mean_list)/ len(average_win_mean_list),3)
    except:
        average_win_mean = None
    output_dict['average_win_mean'] = average_win_mean
    try: 
        average_loss_mean_list = [random_metrics_iter[i]['average_loss'] for i in random_metrics_iter]
        average_loss_mean = round(sum(average_loss_mean_list)/ len(average_loss_mean_list),3)
    except:
        average_loss_mean = None
    output_dict['average_loss_mean'] = average_loss_mean
    try: 
        average_holding_period_mean_list = [random_metrics_iter[i]['average_holding_period'] for i in random_metrics_iter]
        average_holding_period_mean = round(sum(average_holding_period_mean_list)/ len(average_holding_period_mean_list),3)
    except:
        average_holding_period_mean = None
    output_dict['average_holding_period_mean'] = average_holding_period_mean 
    try: 
        max_drawdown_list = [random_metrics_iter[i]['max_drawdown'] for i in random_metrics_iter]
        max_drawdown_mean = round(sum(max_drawdown_list)/ len(max_drawdown_list),3) 
    except:
        max_drawdown_mean = None
    output_dict['max_drawdown_mean'] = max_drawdown_mean
    try: 
        sharpe_ratio_list = [random_metrics_iter[i]['sharpe_ratio'] for i in random_metrics_iter]
        sharpe_ratio_mean = round(sum(sharpe_ratio_list)/ len(sharpe_ratio_list),3)
    except:
        sharpe_ratio_mean = None
    output_dict['sharpe_ratio_mean'] = sharpe_ratio_mean
    try: 
        calmar_ratio_list = [random_metrics_iter[i]['calmar_ratio'] for i in random_metrics_iter]
        calmar_ratio_mean = round(sum(calmar_ratio_list)/ len(calmar_ratio_list),3)
    except:
        calmar_ratio_mean = None
    output_dict['calmar_ratio_mean'] = calmar_ratio_mean
    try:
        sortino_ratio_list = [random_metrics_iter[i]['sortino_ratio'] for i in random_metrics_iter]
        sortino_ratio_mean = round(sum(sortino_ratio_list)/ len(sortino_ratio_list),3)
    except: 
        sortino_ratio_mean = None
    output_dict['sortino_ratio_mean'] = sortino_ratio_mean
    try: 
        omega_ratio_list = [random_metrics_iter[i]['omega_ratio'] for i in random_metrics_iter]
        omega_ratio_mean = round(sum(omega_ratio_list)/ len(omega_ratio_list),3)
    except: 
        omega_ratio_mean = None
    output_dict['omega_ratio_mean'] = omega_ratio_mean
    try: 
        serenity_index_list = [random_metrics_iter[i]['serenity_index'] for i in random_metrics_iter]
        serenity_index_mean = round(sum(serenity_index_list)/ len(serenity_index_list),3)
    except: 
        serenity_index_mean = None
    output_dict['serenity_index_mean'] = serenity_index_mean
    try: 
        smart_sharpe_list = [random_metrics_iter[i]['smart_sharpe'] for i in random_metrics_iter]
        smart_sharpe_mean = round(sum(smart_sharpe_list)/ len(smart_sharpe_list),3)
    except:
        smart_sharpe_mean = None
    output_dict['smart_sharpe_mean'] = smart_sharpe_mean
    try: 
        smart_sortino_list = [random_metrics_iter[i]['smart_sortino'] for i in random_metrics_iter]
        smart_sortino_mean = round(sum(smart_sortino_list)/ len(smart_sortino_list),3)
    except:
        smart_sortino_mean = None
    output_dict['smart_sortino_mean'] = smart_sortino_mean
    try: 
        winning_streak_list = [random_metrics_iter[i]['winning_streak'] for i in random_metrics_iter]
        winning_streak_mean = round(sum(winning_streak_list)/ len(winning_streak_list),3)
    except:
        winning_streak_mean = None
    output_dict['winning_streak_mean'] = winning_streak_mean
    try: 
        losing_streak_list = [random_metrics_iter[i]['losing_streak'] for i in random_metrics_iter]
        losing_streak_mean = round(sum(losing_streak_list)/ len(losing_streak_list),3)
    except:
        losing_streak_mean = None
    output_dict['losing_streak_mean'] = losing_streak_mean
    try: 
        largest_losing_trade_list = [random_metrics_iter[i]['largest_losing_trade'] for i in random_metrics_iter]
        largest_losing_trade_mean = round(sum(largest_losing_trade_list)/ len(largest_losing_trade_list),3)
    except:
        largest_losing_trade_mean = None
    output_dict['largest_losing_trade_mean'] = largest_losing_trade_mean
    try: 
        largest_winning_trade_list = [random_metrics_iter[i]['largest_winning_trade'] for i in random_metrics_iter]
        largest_winning_trade_mean = round(sum(largest_winning_trade_list)/ len(largest_winning_trade_list),3)
    except:
        largest_winning_trade_mean = None
    output_dict['largest_winning_trade_mean'] = largest_winning_trade_mean
    try: 
        kelly_criterion_list = [random_metrics_iter[i]['kelly_criterion'] for i in random_metrics_iter]
        kelly_criterion_mean = round(sum(kelly_criterion_list)/ len(kelly_criterion_list),3)
    except:
        kelly_criterion_mean = None
    output_dict['kelly_criterion_mean'] = kelly_criterion_mean
    output_dict['average_param_%deviation'] = list_rand_average     
    output_dict['start_date'] = robust_start
    output_dict['end_date'] = robust_end
    random_params_list = [random_metrics_iter[i]['random_params'] for i in random_metrics_iter]
    param_len = (random_params_list[0])
    for k in param_len.keys():
        key = k
        output_dict[f'{key}_avg'] = avrg(random_params_list,key) 
    random_df = pd.DataFrame.from_dict(output_dict.items(), orient='columns')
    for i in range(random_df.shape[0]):
        if i == 0:
            new_dataframe.iat[48,15] = f"Average Robustness Testing Results over {robust_length}"
        if i == 1:
            new_dataframe.iat[49,15] = f"Iterations with a Parameter Deviation Range"
        if i == 2:
            new_dataframe.iat[50,15] = f"of {robust_min/100}% to {robust_max/100}% from {robust_start} to {robust_end}"
        for j in range(2):
            new_dataframe.iat[(51+i),(j+15)] = random_df.iat[(0+i),(j)] 
    df__net_profit_p = net_profit_percentage_list
    
    # Changed from two columns of data to one column of data due to introduction of param dev data. 
    
    # if cfg['robust_test_iteration_count'] < 26:
    for i in range(len(df__net_profit_p)):
        if i == 0 :
            new_dataframe.iat[81,15] = "Param % Deviation"
            new_dataframe.iat[81,16] = "Net Profit % per Iteration"
            new_dataframe.iat[81,17] = "Parameters"
        # if i < 13:
        new_dataframe.iat[82+i,15] = final_rand_dev[i]
        new_dataframe.iat[82+i,16] = df__net_profit_p[i]
        new_dataframe.iat[82+i,17] = random_params_list[i]
            # else:
                # new_dataframe.iat[82+(i-12),16] = df__net_profit_p[i]
                # new_dataframe.iat[82+(i-12),17] = random_params_list[i]
    # else:
        # for i in range(len(df__net_profit_p)):
            # if i == 0 :
                # new_dataframe.iat[96,0] = "Robust Net Profit % per Iteration"
            # new_dataframe.iat[97,0] = df__net_profit_p[i] 
            
    new_dataframe.to_pickle(f'{filename}-pickled', protocol=-1)

def random_pairs_func(dir1,output_df,cfg,index,sorted_candidate_trial_list):
    for f in os.listdir(dir1):
        filename_str = f.split(':')[1]
        trial_num_str = str(filename_str.split('-')[0])
        if trial_num_str == str(output_df.iloc[index,0]):
            filename = f"{dir1}{f}"    
            break
    for trial in sorted_candidate_trial_list:
        if str(trial.number) == str(trial_num_str):
            params = trial.params

    index = output_df[output_df.iloc[:,0].str.contains(trial_num_str)].index.values
    
    new_dataframe = pd.read_pickle(f'{filename}') 
    
    crypto_pair_list = ['LTC-USD', 'ETH-USD', 'XMR-USD', 'BTC-USD'] #['LTC-USD','ETH-USD', 'XRP-USD', 'XMR-USD','ETC-USD','ZEC-BTC', 'BTC-USD']
    stock_pair_list = ['NVDA-USD', 'AMD-USD', 'AAPL-USD', 'AMZN-USD', 'TGT-USD', 'F-USD', 'T-USD', 'TSLA-USD', 'CCL-USD', 'INTC-USD', 'GOOG-USD', 'VZ-USD'] 
    symbol_list = []
    other_pair_metrics = {}
    pair_start = cfg['pair-testing']['start_date']
    pair_end = cfg['pair-testing']['finish_date']
    old_pair = cfg['route'][0]['symbol']
    old_exchange = cfg['route'][0]['exchange']
    temp_df_list = []
    if cfg['route'][0]['exchange'] != 'Polygon_Stocks':
        symbol_list = crypto_pair_list
    else: 
        for i in range(4):
            symbol_list.append(random.choice(stock_pair_list))
    if old_exchange != 'Polygon_Stocks':
        cfg['route'][0]['exchange'] = 'Bitfinex Spot'
    if len(cfg['route']) > 1:
        cfg['route'] = {0: {'exchange': cfg['route'][0]['exchange'], 'symbol': old_pair, 'timeframe': cfg['route'][0]['timeframe']} }  
    for index, symbol in enumerate(symbol_list):
        try:
            temp = {}
            cfg['route'][0]['symbol'] = symbol
            temp = optuna_backtest_function(pair_start,pair_end,params,cfg, optimizing=True)
            temp = {k: v if v is not (None or '') else 0 for k, v in temp.items()}
            temp['index'] = index+1
            temp = {key : round(temp[key],3) for key in temp}
            temp['symbol'] = symbol
            temp_df = pd.DataFrame.from_dict(temp.items(), orient='columns')
            temp_df.fillna(0)
        except Exception as e:
            print(e)
            temp_df = temp_df_list[index-1]
            pass
        temp_df_list.append(temp_df)
        
    glob_path = f"{dir1}random_pair*"
    # t_files = glob.glob(glob_path)   
    t_df = pd.concat(temp_df_list,ignore_index=True,axis=1)
    drop_idx = list(range(2,t_df.shape[1],2))
    drop_cols =  [j for i,j in enumerate(t_df.columns) if i in drop_idx]
    t_df = t_df.drop(drop_cols,axis=1) 
    # files_to_delete_path = f"{dir1}random_pair*"
    # files_to_delete = glob.glob(files_to_delete_path)
    # for file in files_to_delete:
        # os.remove(file)
    for i in range(20):
        new_dataframe = new_dataframe.append(pd.Series(), ignore_index = True)
        empty_series = pd.Series()
        new_dataframe = pd.concat([new_dataframe, empty_series],axis=1)
    for i in range(t_df.shape[0]):
        if i == 0:
            new_dataframe.iat[48,19] = f"Backtest Metrics for 4 Random Pairs From"
            new_dataframe.iat[49,19] = f"{pair_start} to {pair_end} on {cfg['route'][0]['timeframe']} Candles"
        for j in range(t_df.shape[1]):
            new_dataframe.iat[50,(j+19)] = t_df.iat[43,(j)]
            new_dataframe.iat[51+i,(19+j)] = t_df.iat[(0+i),(j)]

    new_dataframe.to_pickle(f'{filename}')
    cfg['route'][0]['exchange'] = old_exchange
    cfg['route'][0]['symbol'] = old_pair
            
def metrics_func(best_trial,missing_files,extra_files,cfg,Interval_start,Interval_end,dir1):
    try:
        final_path = f"{dir1}final-trial:{best_trial.number}-{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-interval:{cfg['validation_interval']}-{Interval_start}-{Interval_end}.csv"
        formated = f"final-trial:{best_trial.number}-{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-interval:{cfg['validation_interval']}-{Interval_start}-{Interval_end}.xlsx"      
        _formated_final_path = "unknown.csv"
        # Check to see if the trial number has fully finished and formatted into an .xlsx file. 
        for f in os.listdir(dir1):
            if f.endswith(formated):
                _formated_final_path = f         
        formated_final_path = f"{dir1}{_formated_final_path}"
        # rerun if incomplete or missing
        if not os.path.isfile(formated_final_path) and not os.path.isfile(final_path):
            begin = datetime.strptime(str(Interval_start), '%Y-%m-%d')
            end = datetime.strptime(str(Interval_end), '%Y-%m-%d')
            interval = cfg['validation_interval'] # interval in days 
            date_list =  list(map(str, date_groups(begin, end, interval))) 
            for index, item in enumerate(date_list[:-1]):
                begin_string = f'{item[:4]}-{item[5:7]}-{item[8:10]}'
                next_index = date_list[index+1]
                end_string = f'{next_index[:4]}-{next_index[5:7]}-{next_index[8:10]}'
                # print(f' Item: {item}')
                # print(f' Next: {date_list[index+1]}')
                validation_data = optuna_backtest_function(item,date_list[index+1],best_trial.params,cfg, optimizing=True)
                validation_data = {k: v if v is not (None or '' or v != v) else 0 for k, v in validation_data.items()}
                try:
                    validation_data = {key : round(validation_data[key],3) for key in validation_data}
                except Exception as e:
                    print(e)
                    pass
                validation_data['start_date'] = begin_string
                validation_data['index'] = (index+1)
                if index == 0:
                    validation_data['parameters'] = best_trial.params           
                validation_df = pd.DataFrame.from_dict(validation_data.items(), orient='columns')
                validation_df.fillna(0)
                validation_df.to_csv(f"{dir1}temp/trial:{best_trial.number}-{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-interval:{cfg['validation_interval']}-index:{index}-{begin_string}-{end_string}.csv", header=True, index=False, encoding='utf-8', sep=',')
            glob_path = f"{dir1}temp/trial:{best_trial.number}-{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-interval:{cfg['validation_interval']}*.csv"
            files = glob.glob(glob_path)
            # Check to see if the interval backtesting has been completed
            try: 
                final_df = pd.concat(map(pd.read_csv, files), ignore_index=True, axis=1)
            except:
                print(f'{Fore.RED} Possibly Missing Files in Temp Folder {Style.RESET_ALL}')
                pass
            drop_idx = list(range(2,final_df.shape[1],2))
            drop_cols =  [j for i,j in enumerate(final_df.columns) if i in drop_idx]
            final_df = final_df.drop(drop_cols,axis=1)
            final_df = final_df.sort_values(by = 42, ascending=True, axis=1)
            cols = list(final_df.columns)
            cols = [cols[-1]] + cols[:-1]
            final_df = final_df[cols]
            final_df = final_df.T.drop_duplicates().T
            final_df.to_csv(final_path,header=False, index=False, encoding='utf-8', sep=',') 
            files_to_delete_path = f"{dir1}/temp/trial:{best_trial.number}-{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-interval:{cfg['validation_interval']}*.csv"
            files_to_delete = glob.glob(files_to_delete_path)
            # delete temporary files which held the pandas dataframe information
            for file in files_to_delete:
                try:
                    if file != f'removed_candidates.pkl':
                        os.remove(file)
                except:
                    extra_files = True
                    continue
            missing_files = True
        return (missing_files,extra_files)
    except Exception as e:
        print(e)
        files_to_delete_path = f"{dir1}/temp/trial:{best_trial.number}-{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-interval:{cfg['validation_interval']}*.csv"
        files_to_delete = glob.glob(files_to_delete_path)
        # delete temporary files which held the pandas dataframe information
        for file in files_to_delete:
            try:
                if file != f'removed_candidates.pkl':
                    os.remove(file)
            except:
                extra_files = True
                continue
        with open(f'{dir1}/{best_trial.number}-delete', 'w') as file:
            pass 
        pass
    
def color_negative(v,color):
    return f"color: {color};" if (v is int or v is float) and v < 0 else None
    
def color_positive(v,color):
    return f"color: {color};" if (v is int or v is float) and v > 0 else None
    
def bold_font(v):
    return "font-weight: bold;" if v is str else None
    
def date_groups( 
    start_at: datetime, 
    end_at: datetime,
    max_capacity_days: float) -> Iterable[datetime]:
    capacity = timedelta(days=max_capacity_days)
    interval = int( ((end_at)  - start_at ) / capacity) + 1
    for i in range(interval):
        step_date = (start_at + capacity * i)
        if ((datetime.strptime(str(end_at), "%Y-%m-%d %H:%M:%S") - datetime.strptime(str(step_date), "%Y-%m-%d %H:%M:%S")).days) > 30:
            yield step_date
    yield end_at
    
def get_config():
    cfg_file = pathlib.Path('optuna_config.yml')

    if not cfg_file.is_file():
        print("optuna_config.yml not found. Run create-config command.")
        exit()
    else:
        with open("optuna_config.yml", "r") as ymlfile:
            cfg = yaml.load(ymlfile, yaml.SafeLoader)

    return cfg

def write_dict_to_yaml(input_dict, filename="optuna_config.yml"):
    with open(filename, 'w') as file:
        yaml.dump(input_dict, file, default_flow_style=False)

def _check_termination():
    return os.path.exists("./storage/temp/termination_signal.txt")
        
def objective(trial):
    start_time = time.time()
    if _check_termination():
        trial.study.stop()

    cfg = get_config()
    resumed = False
    StrategyClass = jh.get_strategy_class(cfg['strategy_name'])
    hp_dict = StrategyClass().hyperparameters()
    fitness_description = f"{cfg['fitness-ratio1']} - {cfg['fitness-ratio2']}" if cfg['mode'] == 'multi' else f"{cfg['fitness-ratio1']}"
    # load previous trial data
    if (trial.number - (cfg['n_trials_start']+1) == 0) and trial.number > 50:
        resumed = True
        storage = f"postgresql://jesse_user:{cfg['postgres_password']}@{cfg['postgres_host']}:{cfg['postgres_port']}/optuna_db"
        study_history = optuna.load_study(study_name=cfg['study_full_name'], storage=storage)
        best_candidates_array = []
        all_trials =  study_history.get_trials(states=[optuna.trial.TrialState.COMPLETE])
        if len(all_trials) > 0:
            for _trial in all_trials:
                dna_string = jh.hp_to_dna(hp_dict, _trial.params)
                best_candidate = {
                    'rank': _trial.number,
                    'dna': dna_string,
                    'fitness': round(_trial.user_attrs.get('training_score_1',-1),3) if cfg['mode'] == 'single' else (round(_trial.user_attrs.get('training_score_1',-1),3),round(_trial.user_attrs.get('training_score_2',-1),3)),  
                    'training_win_rate': int(_trial.user_attrs.get('training-win_rate',-1) * 100),
                    'training_total_trades': _trial.user_attrs.get('training-total',-1),
                    'training_pnl': round(_trial.user_attrs.get('training-net_profit_percentage',-1), 2),
                    'testing_win_rate': int(_trial.user_attrs.get('testing-win_rate',-1) * 100),
                    'testing_total_trades': _trial.user_attrs.get('testing-total',-1),
                    'testing_pnl': round(_trial.user_attrs.get('testing-net_profit_percentage',-1),2),
                    'hyperparameters': _trial.params,
                    'fitness_description': fitness_description
                }
                best_candidates_array.append(best_candidate)
            sync_publish('best_candidates', best_candidates_array, 'optuna', cfg['main_pid'])
    # initialize default values if the trial is pruned early.
    if cfg['optimizer'] in ['NSGAIISampler', 'TPESampler']:
        default_constraints = (0.5, 0.5)  
        trial.set_user_attr("constraint", default_constraints)

    # Define hyperparameters based on the sampler type
    if cfg['optimizer'] == 'GridSampler':
        # Just validate the names and types; the values are fixed by the grid
        for st_hp in hp_dict:
            if st_hp['type'] is int:
                trial.suggest_int(st_hp['name'], st_hp['min'], st_hp['max'])
            elif st_hp['type'] is float:
                trial.suggest_float(st_hp['name'], st_hp['min'], st_hp['max'])
            elif st_hp['type'] is bool:
                trial.suggest_categorical(st_hp['name'], [True, False])
            elif st_hp['type'] is str:
                trial.suggest_categorical(st_hp['name'], st_hp['list'])
    else:
        # Dynamically suggest values
        for st_hp in hp_dict:
            if st_hp['type'] is int:
                step = st_hp.get('step', 1)
                trial.suggest_int(st_hp['name'], st_hp['min'], st_hp['max'], step=step)
            elif st_hp['type'] is float:
                step = st_hp.get('step', 0.1)
                trial.suggest_float(st_hp['name'], st_hp['min'], st_hp['max'], step=step)
            elif st_hp['type'] is bool:
                trial.suggest_categorical(st_hp['name'], [True, False])
            elif st_hp['type'] is str:
                trial.suggest_categorical(st_hp['name'], st_hp['list'])

    try:
        training_data_metrics = optuna_backtest_function(cfg['timespan-train']['start_date'],
                                                  cfg['timespan-train']['finish_date'],
                                                  trial.params, cfg, optimizing=True)
        if training_data_metrics is None:
            raise optuna.TrialPruned()
    except Exception as err:
        logger.error("".join(traceback.TracebackException.from_exception(err).format()))
        raise err
        
    if training_data_metrics['total'] <= cfg['optimal-total']:
        raise optuna.TrialPruned()

    total_effect_rate = np.log10(training_data_metrics['total']) / np.log10(cfg['optimal-total'])
    total_effect_rate = min(total_effect_rate, 1)
    ratio_config = cfg['fitness-ratio1']
    if ratio_config == 'Sharpe':
        ratio = training_data_metrics['sharpe_ratio']
        ratio_normalized = jh.normalize(ratio, -.5, 5)
    elif ratio_config == 'Calmar':
        ratio = training_data_metrics['calmar_ratio']
        ratio_normalized = jh.normalize(ratio, -.5, 30)
    elif ratio_config == 'Sortino':
        ratio = training_data_metrics['sortino_ratio']
        ratio_normalized = jh.normalize(ratio, -.5, 15)
    elif ratio_config == 'Omega':
        ratio = training_data_metrics['omega_ratio']
        ratio_normalized = jh.normalize(ratio, -.5, 5)
    elif ratio_config == 'Serenity':
        ratio = training_data_metrics['serenity_index']
        ratio_normalized = jh.normalize(ratio, -.5, 15)
    elif ratio_config == 'SmartSharpe':
        ratio = training_data_metrics['smart_sharpe']
        ratio_normalized = jh.normalize(ratio, -.5, 5)
    elif ratio_config == 'SmartSortino':
        ratio = training_data_metrics['smart_sortino']
        ratio_normalized = jh.normalize(ratio, -.5, 15)
    elif ratio_config == 'MaxDrawdown':
        ratio = training_data_metrics['max_drawdown']
        ratio_normalized = ratio 
    elif ratio_config == 'GrossLoss':
        ratio = training_data_metrics['gross_loss']
        ratio_normalized = ratio
    else:
        raise ValueError(
            f'The entered ratio configuration `{ratio_config}` for the optimization is unknown. Choose between Sharpe, Calmar, Sortino, Serenity, SmartSharpe, SmartSortino and Omega.')
    if ratio < 0:
        raise optuna.TrialPruned

    score = total_effect_rate * ratio_normalized
    if cfg['mode'] == 'single':
        trial.report(score, 1)
    trial.set_user_attr(f"training_score_1", score)
    
    if cfg['mode'] == 'multi':
        ratio_config2 = cfg['fitness-ratio2']
        if ratio_config2 == 'Sharpe':
            ratio2 = training_data_metrics['sharpe_ratio']
            ratio_normalized2 = jh.normalize(ratio2, -.5, 5)
        elif ratio_config2 == 'Calmar':
            ratio2 = training_data_metrics['calmar_ratio']
            ratio_normalized2 = jh.normalize(ratio2, -.5, 30)
        elif ratio_config2 == 'Sortino':
            ratio2 = training_data_metrics['sortino_ratio']
            ratio_normalized2 = jh.normalize(ratio2, -.5, 15)
        elif ratio_config2 == 'Omega':
            ratio2 = training_data_metrics['omega_ratio']
            ratio_normalized2 = jh.normalize(ratio2, -.5, 5)
        elif ratio_config2 == 'Serenity':
            ratio2 = training_data_metrics['serenity_index']
            ratio_normalized2 = jh.normalize(ratio2, -.5, 15)
        elif ratio_config2 == 'SmartSharpe':
            ratio2 = training_data_metrics['smart_sharpe']
            ratio_normalized2 = jh.normalize(ratio2, -.5, 5)
        elif ratio_config2 == 'SmartSortino':
            ratio2 = training_data_metrics['smart_sortino']
            ratio_normalized2 = jh.normalize(ratio2, -.5, 15)
        elif ratio_config2 == 'MaxDrawdown':
            ratio2 = training_data_metrics['max_drawdown']
            ratio_normalized2 = ratio2
        elif ratio_config2 == 'GrossLoss':
            ratio2 = training_data_metrics['gross_loss']
            ratio_normalized2 = ratio2
        else:
            raise ValueError(
                f'The entered ratio2 configuration `{ratio_config2}` for the optimization is unknown. Choose between Sharpe, Calmar, Sortino, Serenity, smart shapre, SmartSortino and Omega.')
        if ratio2 < 0 and cfg['dual_mode'] == 'maximize':
            raise optuna.TrialPruned()
        score2 = total_effect_rate * ratio_normalized2
        trial.set_user_attr(f"training_score_2", score2)
    
    try:
        testing_data_metrics = optuna_backtest_function(cfg['timespan-testing']['start_date'], cfg['timespan-testing']['finish_date'], trial.params, cfg, optimizing=True)
        if testing_data_metrics is None:
            raise optuna.TrialPruned()
    except Exception as err:
        logger.error("".join(traceback.TracebackException.from_exception(err).format()))
        raise err
        
    #constraints_violation must be negative to proceed. Only used for NSGAIISampler and TPESampler
    if cfg['optimizer'] in ['NSGAIISampler', 'TPESampler']:
        constraints_violation1 = (-1*testing_data_metrics['omega_ratio'])+0.85
        constraints_violation2 = -1*testing_data_metrics['kelly_criterion']
        trial.set_user_attr("constraint", (constraints_violation1,constraints_violation2))
        # print(constraints_violation1)
        # print(constraints_violation2)
    
    for key, value in testing_data_metrics.items():
        if isinstance(value, np.integer):
            value = int(value)
        elif isinstance(value, np.floating):
            value = float(value)
        elif isinstance(value, np.ndarray):
            value = value.tolist()
        trial.set_user_attr(f"testing-{key}", value)

    for key, value in training_data_metrics.items():
        if isinstance(value, np.integer):
            value = int(value)
        elif isinstance(value, np.floating):
            value = float(value)
        elif isinstance(value, np.ndarray):
            value = value.tolist()
        trial.set_user_attr(f"training-{key}", value) 
        
    testing_total_effect_rate = np.log10(testing_data_metrics['total']) / np.log10(cfg['optimal-total'])
    testing_total_effect_rate = min(testing_total_effect_rate, 1) 
    
    ratio_config = cfg['fitness-ratio1']
    if ratio_config == 'Sharpe':
        testing_ratio = testing_data_metrics['sharpe_ratio']
        testing_ratio_normalized = jh.normalize(testing_ratio, -.5, 5)
    elif ratio_config == 'Calmar':
        testing_ratio = testing_data_metrics['calmar_ratio']
        testing_ratio_normalized = jh.normalize(testing_ratio, -.5, 30)
    elif ratio_config == 'Sortino':
        testing_ratio = testing_data_metrics['sortino_ratio']
        testing_ratio_normalized = jh.normalize(testing_ratio, -.5, 15)
    elif ratio_config == 'Omega':
        testing_ratio = testing_data_metrics['omega_ratio']
        testing_ratio_normalized = jh.normalize(testing_ratio, -.5, 5)
    elif ratio_config == 'Serenity':
        testing_ratio = testing_data_metrics['serenity_index']
        testing_ratio_normalized = jh.normalize(testing_ratio, -.5, 15)
    elif ratio_config == 'SmartSharpe':
        testing_ratio = testing_data_metrics['smart_sharpe']
        testing_ratio_normalized = jh.normalize(testing_ratio, -.5, 5)
    elif ratio_config == 'SmartSortino':
        testing_ratio = testing_data_metrics['smart_sortino']
        testing_ratio_normalized = jh.normalize(testing_ratio, -.5, 15)
    elif ratio_config == 'MaxDrawdown':
        testing_ratio = testing_data_metrics['max_drawdown']
        testing_ratio_normalized = testing_ratio
    elif ratio_config == 'GrossLoss':
        testing_ratio = testing_data_metrics['gross_loss']
        testing_ratio_normalized = testing_ratio
    else:
        raise ValueError(
            f'The entered ratio configuration `{ratio_config}` for the optimization is unknown. Choose between Sharpe, Calmar, Sortino, Serenity, smart shapre, SmartSortino and Omega.')
    testing_score = testing_total_effect_rate * testing_ratio_normalized
    trial.set_user_attr(f"testing_score_1", testing_score)   
    
    if testing_ratio_normalized < (-(ratio_normalized/7)):
        raise optuna.TrialPruned()     
        
    if cfg['mode'] == 'multi':
        ratio_config2 = cfg['fitness-ratio2']
        if ratio_config2 == 'Sharpe':
            testing_ratio2 = testing_data_metrics['sharpe_ratio']
            testing_ratio_normalized2 = jh.normalize(testing_ratio2, -.5, 5)
        elif ratio_config2 == 'Calmar':
            testing_ratio2 = testing_data_metrics['calmar_ratio']
            testing_ratio_normalized2 = jh.normalize(testing_ratio2, -.5, 30)
        elif ratio_config2 == 'Sortino':
            testing_ratio2 = testing_data_metrics['sortino_ratio']
            testing_ratio_normalized2 = jh.normalize(testing_ratio2, -.5, 15)
        elif ratio_config2 == 'Omega':
            testing_ratio2 = testing_data_metrics['omega_ratio']
            testing_ratio_normalized2 = jh.normalize(testing_ratio2, -.5, 5)
        elif ratio_config2 == 'Serenity':
            testing_ratio2 = testing_data_metrics['serenity_index']
            testing_ratio_normalized2 = jh.normalize(testing_ratio2, -.5, 15)
        elif ratio_config2 == 'SmartSharpe':
            testing_ratio2 = testing_data_metrics['smart_sharpe']
            testing_ratio_normalized2 = jh.normalize(testing_ratio2, -.5, 5)
        elif ratio_config2 == 'SmartSortino':
            testing_ratio2 = testing_data_metrics['smart_sortino']
            testing_ratio_normalized2 = jh.normalize(testing_ratio2, -.5, 15)
        elif ratio_config2 == 'MaxDrawdown':
            testing_ratio2 = testing_data_metrics['max_drawdown']
            testing_ratio_normalized2 = testing_ratio2
        elif ratio_config2 == 'GrossLoss':
            testing_ratio2 = testing_data_metrics['gross_loss']
            testing_ratio_normalized2 = testing_ratio2
        else:
            raise ValueError(
                f'The entered ratio2 configuration `{ratio_config2}` for the optimization is unknown. Choose between Sharpe, Calmar, Sortino, Serenity, smart shapre, SmartSortino and Omega.')
        if testing_ratio_normalized2 < (-(ratio_normalized2/5)) and cfg['dual_mode'] == 'maximize':
            raise optuna.TrialPruned()
        testing_score_2 = testing_total_effect_rate * testing_ratio_normalized2
        trial.set_user_attr(f"testing_score_2", testing_score_2)
     
     
    if cfg['debug_mode']:
        training_data_metrics['start_date'] = cfg['timespan-train']['start_date']
        training_data_metrics['finish_date'] = cfg['timespan-train']['finish_date']
        training_data_metrics['parameters'] = trial.params
        pairs = []
        for i in range(len(cfg['route'].items())):
            pairs.append(cfg['route'][i]['symbol'])
        pairs = ','.join(pairs)
        training_data_metrics['pairs'] = pairs
        training_results_df = pd.DataFrame.from_dict(training_data_metrics.items(), orient='columns')
        testing_data_metrics['start_date'] = cfg['timespan-testing']['start_date']
        testing_data_metrics['finish_date'] = cfg['timespan-testing']['finish_date']
        testing_data_metrics['parameters'] = trial.params
        testing_data_metrics['pairs'] = pairs
        testing_results_df = pd.DataFrame.from_dict(testing_data_metrics.items(), orient='columns')  
        train_temp_path = f"./storage/jesse-optuna/training_metrics/"
        test_temp_path = f"./storage/jesse-optuna/testing_metrics/"
        train_path = f"{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['type']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{cfg['timespan-train']['start_date']}-{cfg['timespan-train']['finish_date']}-{cfg['optimizer']}-{len(cfg['route'])} Pair"
        test_path = f"{cfg['strategy_name']}-{cfg['route'][0]['exchange']}-{cfg['type']}-{cfg['route'][0]['symbol']}-{cfg['route'][0]['timeframe']}-{cfg['timespan-train']['start_date']}-{cfg['timespan-train']['finish_date']}-{cfg['optimizer']}-{len(cfg['route'])} Pair"
        
        if not os.path.exists(f"{train_temp_path}{train_path}"):
            os.makedirs(f"{train_temp_path}{train_path}")
        if not os.path.exists(f"{test_temp_path}{test_path}"):
            os.makedirs(f"{test_temp_path}{test_path}")
        training_results_df.to_csv(f"{train_temp_path}{train_path}/{round(training_data_metrics['smart_sharpe'],3)} sharpie -training: {trial.number}-{cfg['strategy_name']}.csv", header=True, index=False, encoding='utf-8', sep=',')
        testing_results_df.to_csv(f"{test_temp_path}{test_path}/{round(testing_data_metrics['smart_sharpe'],3)} sharpie -testing: {trial.number}-{cfg['strategy_name']}.csv", header=True, index=False, encoding='utf-8', sep=',')
    
    duration = round(time.time() - start_time,3)
    progressbar_current = int((trial.number - cfg['n_trials_start']) / cfg['n_trials'] * 100)
    sync_publish('progressbar', {
        'current': progressbar_current,
        'estimated_remaining_seconds': int((duration*((cfg['n_trials'] + cfg['n_trials_start']) - trial.number))/cfg['n_jobs'])
    },'optuna',cfg['main_pid'])
    
    dna_string = (jh.hp_to_dna(hp_dict,trial.params))
        
    if not resumed:
        best_candidates = {
                'rank': '',
                'dna': dna_string,
                'fitness' : round(score,3) if cfg['mode'] == 'single' else (round(score,3),round(score2,3)),
                'training_win_rate': int(training_data_metrics['win_rate'] * 100),
                'training_total_trades': training_data_metrics['total'],
                'training_pnl': round(training_data_metrics['net_profit_percentage'], 2),
                'testing_win_rate': int(testing_data_metrics['win_rate'] * 100),
                'testing_total_trades': testing_data_metrics['total'],
                'testing_pnl': round(testing_data_metrics['net_profit_percentage'],2),
                'hyperparameters': trial.params,
                'fitness_description': fitness_description
            }
        sync_publish('best_candidates', best_candidates, 'optuna', cfg['main_pid'])

    general_info = {
        'started_at': jh.timestamp_to_arrow(cfg['start_time']).humanize(),
        'index': f"{trial.number}/{cfg['n_trials']+cfg['n_trials_start']}",
        'trading_route': '-'.join(cfg['study_full_name'].split('-')[0:-1]),
        'average_execution_seconds': duration,
        'optimizer': cfg['optimizer']
        }
    sync_publish('general_info', general_info,'optuna',cfg['main_pid'])
    if cfg['mode'] == 'single':
        return score
    else:
        return score, score2

def constraints_function(trial):
    return trial.user_attrs["constraint"]
  
def validate_cwd() -> None:
    """
    make sure we're in a Jesse project
    """
    ls = os.listdir('.')
    is_jesse_project = 'strategies' in ls and 'storage' in ls

    if not is_jesse_project:
        print('Current directory is not a Jesse project. You must run commands from the root of a Jesse project.')
        exit()


def get_candles_with_cache(exchange: str, symbol: str, start_date: str, finish_date: str) -> np.ndarray:
    path = pathlib.Path('storage/jesse-optuna/candle_storage/')
    path.mkdir(parents=True, exist_ok=True)

    cache_file_name = f"{exchange}-{symbol}-1m-{start_date}-{finish_date}.pickle"
    cache_file = pathlib.Path(f'storage/jesse-optuna/candle_storage/{cache_file_name}')

    if cache_file.is_file():
        with open(f'storage/jesse-optuna/candle_storage/{cache_file_name}', 'rb') as handle:
            candles = pickle.load(handle)
    else:
        candles = get_candles(exchange, symbol, '1m', start_date, finish_date)
        with open(f'storage/jesse-optuna/candle_storage/{cache_file_name}', 'wb') as handle:
            pickle.dump(candles, handle, protocol=pickle.HIGHEST_PROTOCOL)
    
    return candles

@memory.cache
def optuna_backtest_function(start_date, finish_date, hp, cfg, run_silently=True,trial_num=None, optimizing=None):
    candles = {}
    extra_routes = []
    route = []
    if len(cfg['route'].items()) > 1:
        for route_ in cfg['route'].items():
            route_ = route_[1]
            candles[jh.key(route_['exchange'], route_['symbol'])] = {
                'exchange': route_['exchange'],
                'symbol': route_['symbol'],
                'candles': get_candles_with_cache(
                    route_['exchange'],
                    route_['symbol'],
                    start_date,
                    finish_date,
                ),
            }
            route.append({'exchange': route_['exchange'], 'strategy': cfg['strategy_name'], 'symbol': route_['symbol'],
                                 'timeframe': route_['timeframe']})
    else:
        candles[jh.key(cfg['route'][0]['exchange'], cfg['route'][0]['symbol'])] = {
        'exchange': cfg['route'][0]['exchange'],
        'symbol': cfg['route'][0]['symbol'],
        'candles': get_candles_with_cache(
            cfg['route'][0]['exchange'],
            cfg['route'][0]['symbol'],
            start_date,
            finish_date,
        ),
    }

        route = [{'exchange': cfg['route'][0]['exchange'], 'strategy': cfg['strategy_name'], 'symbol': cfg['route'][0]['symbol'],
              'timeframe': cfg['route'][0]['timeframe']}]


    config = {
        'starting_balance': cfg['starting_balance'],
        'fee': cfg['fee'],
        'type': cfg['type'],
        'futures_leverage': cfg['futures_leverage'],
        'futures_leverage_mode': cfg['futures_leverage_mode'],
        'exchange': cfg['route'][0]['exchange'],
        'symbol': cfg['route'][0]['symbol'],
        'settlement_currency': cfg['settlement_currency'],
        'warm_up_candles': cfg['warm_up_candles'],
        'start_date' : cfg['timespan-train']['start_date'],
        'finish_date' : cfg['Interval_end'],
        'trial_number' : trial_num,
        'extra_route' : None,
    }
    backtest_data = backtest(config, route, extra_routes, candles, run_silently= run_silently , hyperparameters= hp, optimizing=optimizing)
   
    if backtest_data['total'] == 0:
        backtest_data = {'total': 0, 'total_winning_trades': None, 'total_losing_trades': None,
                         'starting_balance': None, 'finishing_balance': None, 'win_rate': None, 'profit_factor': None,
                         'ratio_avg_win_loss': None, 'longs_count': None, 'longs_percentage': None,
                         'shorts_percentage': None, 'shorts_count': None, 'fee': None, 'net_profit': None,
                         'net_profit_percentage': None, 'average_win': None, 'average_loss': None, 'expectancy': None,
                         'expectancy_percentage': None, 'expected_net_profit_every_100_trades': None,
                         'average_holding_period': None, 'average_winning_holding_period': None,
                         'average_losing_holding_period': None, 'gross_profit': None, 'gross_loss': None,
                         'max_drawdown': None, 'annual_return': None, 'sharpe_ratio': None, 'calmar_ratio': None,
                         'sortino_ratio': None, 'omega_ratio': None, 'serenity_index': None, 'smart_sharpe': None,
                         'smart_sortino': None, 'total_open_trades': None, 'open_pl': None, 'winning_streak': None,
                         'losing_streak': None, 'largest_losing_trade': None, 'largest_winning_trade': None,
                         'current_streak': None,'kelly_criterion': None}
    #convert holding periods from seconds to hours 
    else:
        backtest_data['average_holding_period'] = ((backtest_data['average_holding_period']/60)/60) 
        backtest_data['average_winning_holding_period'] =  ((backtest_data['average_winning_holding_period']/60)/60) 
        backtest_data['average_losing_holding_period'] =  ((backtest_data['average_losing_holding_period']/60)/60) 

    return backtest_data   

def optuna_backtest_function_quant(start_date, finish_date, hp, cfg, run_silently:bool=True,trial_num:int=None, imported_string: str = None,additional_files:bool=False, generate_charts=False, generate_equity_curve=False, full_path_name = None, full_name=None):
    candles = {}
    extra_routes = []
    route = []
    if len(cfg['route']) > 1:
        for route_ in cfg['route'].items():
            route_ = route_[1]
            candles[jh.key(route_['exchange'], route_['symbol'])] = {
                'exchange': route_['exchange'],
                'symbol': route_['symbol'],
                'candles': get_candles_with_cache(
                    route_['exchange'],
                    route_['symbol'],
                    start_date,
                    finish_date,
                ),
            }
            route.append({'exchange': route_['exchange'], 'strategy': cfg['strategy_name'], 'symbol': route_['symbol'],
                                 'timeframe': route_['timeframe']})
                                 
    else:
        candles[jh.key(cfg['route'][0]['exchange'], cfg['route'][0]['symbol'])] = {
        'exchange': cfg['route'][0]['exchange'],
        'symbol': cfg['route'][0]['symbol'],
        'candles': get_candles_with_cache(
            cfg['route'][0]['exchange'],
            cfg['route'][0]['symbol'],
            start_date,
            finish_date,
        ),
    }

        route = [{'exchange': cfg['route'][0]['exchange'], 'strategy': cfg['strategy_name'], 'symbol': cfg['route'][0]['symbol'],
                'timeframe': cfg['route'][0]['timeframe']}]

    config = {
        'starting_balance': cfg['starting_balance'],
        'fee': cfg['fee'],
        'type': cfg['type'],
        'futures_leverage': cfg['futures_leverage'],
        'futures_leverage_mode': cfg['futures_leverage_mode'],
        'exchange': cfg['route'][0]['exchange'],
        'symbol': cfg['route'][0]['symbol'],
        'settlement_currency': cfg['settlement_currency'],
        'warm_up_candles': cfg['warm_up_candles'],
        'start_date' : cfg['Interval_start'],
        'finish_date' : cfg['Interval_end'],
        'trial_number' : trial_num,
        'extra_route' : None,
        'strategy_name' : cfg['strategy_name'],
        'timeframe': cfg['route'][0]['timeframe']
    }

    backtest_data = backtest(config, route, extra_routes, candles, generate_charts= generate_charts, generate_equity_curve=generate_equity_curve, run_silently= run_silently , hyperparameters= hp, imported_string= imported_string, full_path_name=full_path_name, full_name=full_name)

    if backtest_data['metrics']['total'] == 0:
        backtest_data['metrics'] = {'total': 0, 'total_winning_trades': None, 'total_losing_trades': None,
                         'starting_balance': None, 'finishing_balance': None, 'win_rate': None, 'profit_factor': None,
                         'ratio_avg_win_loss': None, 'longs_count': None, 'longs_percentage': None,
                         'shorts_percentage': None, 'shorts_count': None, 'fee': None, 'net_profit': None,
                         'net_profit_percentage': None, 'average_win': None, 'average_loss': None, 'expectancy': None,
                         'expectancy_percentage': None, 'expected_net_profit_every_100_trades': None,
                         'average_holding_period': None, 'average_winning_holding_period': None,
                         'average_losing_holding_period': None, 'gross_profit': None, 'gross_loss': None,
                         'max_drawdown': None, 'annual_return': None, 'sharpe_ratio': None, 'calmar_ratio': None,
                         'sortino_ratio': None, 'omega_ratio': None, 'serenity_index': None, 'smart_sharpe': None,
                         'smart_sortino': None, 'total_open_trades': None, 'open_pl': None, 'winning_streak': None,
                         'losing_streak': None, 'largest_losing_trade': None, 'largest_winning_trade': None,
                         'current_streak': None,'kelly_criterion': None}
    #convert holding periods from seconds to hours 
    else:
        backtest_data['metrics']['average_holding_period'] = ((backtest_data['metrics']['average_holding_period']/60)/60) 
        backtest_data['metrics']['average_winning_holding_period'] =  ((backtest_data['metrics']['average_winning_holding_period']/60)/60) 
        backtest_data['metrics']['average_losing_holding_period'] =  ((backtest_data['metrics']['average_losing_holding_period']/60)/60) 
    if additional_files:
        return backtest_data
    else:
        return backtest_data['metrics']
    

    
